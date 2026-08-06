# -*- coding: utf-8 -*-
"""
Tool import bảng giá bơm chữa cháy từ PDF/Excel của nhà cung cấp.

Quy ước giá:
    gia_dau_vao (giá vốn) = giá NCC báo
    gia_ban_le             = gia_dau_vao × markup (mặc định 1.5 = 150%)

Cách dùng:
    python import_bom_catalog.py bao_gia_affetti.pdf --hang AFFETTI --merge
    python import_bom_catalog.py bao_gia_affetti.pdf --hang AFFETTI --markup 1.6
"""
import sys
import os
import re
import json
import argparse
import datetime


SECTION_PATTERNS = [
    (r"I\s+MÁY BƠM ĐIỆN LIỀN TRỤC", "lien_truc_2900"),
    (r"II\s+MÁY BƠM ĐIỆN RỜI TRỤC", "roi_truc_2900"),
    (r"III\s+MÁY BƠM DIESEL", "diesel_3000"),
    (r"IV\s+MÁY BƠM BÙ ÁP", "bu_ap"),
    (r"V\s+MÁY BƠM ĐIỆN ĐA TẦNG", "da_tang"),
]

ROW_PATTERN = re.compile(
    r"(AFT\S+|AF\S+|AVS\S+|AV\S+)\s+"
    r"([\d.,]+)\s*[-–]\s*([\d.,]+)\s+"
    r"([\d.,]+)\s*[-–]\s*([\d.,]+)\s+"
    r"([\d.,]+)"
)

SECTION_LABELS = {
    "lien_truc_2900": "Cụm bơm điện liền trục 2900 RPM",
    "roi_truc_2900": "Cụm bơm điện rời trục 2900 RPM",
    "diesel_3000": "Cụm bơm diesel 3000 RPM",
    "bu_ap": "Bơm bù áp",
    "da_tang": "Bơm điện đa tầng",
}


def parse_pdf(file_path):
    try:
        import pypdf
    except ImportError:
        sys.exit("Can cai: pip install pypdf")
    r = pypdf.PdfReader(file_path)
    full = "\n".join(p.extract_text() or "" for p in r.pages)

    section_regex = re.compile(
        "(" + "|".join(p for p, _ in SECTION_PATTERNS) + ")",
        re.IGNORECASE)
    parts = section_regex.split(full)
    result = {sk: [] for _, sk in SECTION_PATTERNS}

    for i in range(1, len(parts), 2):
        header = parts[i].strip()
        content = parts[i + 1] if i + 1 < len(parts) else ""
        key = None
        for pat, sk in SECTION_PATTERNS:
            if re.search(pat, header, re.IGNORECASE):
                key = sk
                break
        if not key:
            continue
        for m in ROW_PATTERN.finditer(content):
            model, q1, q2, h1, h2, gia = m.groups()
            kw = None
            if "/" in model:
                tail = model.rsplit("/", 1)[1]
                try:
                    kw = float(tail)
                except ValueError:
                    pass
            if kw is None:
                ctx_start = max(0, m.start() - 100)
                ctx = content[ctx_start:m.start()]
                matches = list(re.finditer(r"(\d+(?:\.\d+)?)\s*[Kk][Ww]", ctx))
                if matches:
                    kw = float(matches[-1].group(1))
            result[key].append({
                "model": model.strip(),
                "kw": kw,
                "q_min": float(q1.replace(",", ".")),
                "q_max": float(q2.replace(",", ".")),
                "h_min": min(float(h1.replace(",", ".")),
                             float(h2.replace(",", "."))),
                "h_max": max(float(h1.replace(",", ".")),
                             float(h2.replace(",", "."))),
                "gia": int(gia.replace(",", "").replace(".", "")),
            })
    return result


def parse_xlsx(file_path):
    try:
        import openpyxl
    except ImportError:
        sys.exit("Can cai: pip install openpyxl")
    wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
    result = {}
    for sh in wb.sheetnames:
        ws = wb[sh]
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            continue
        header = [str(c or "").strip().lower() for c in rows[0]]
        idx = {h: i for i, h in enumerate(header)}
        section_key = sh.lower().strip().replace(" ", "_")
        for row in rows[1:]:
            if not row or not row[0]:
                continue
            try:
                model = str(row[idx.get("model", 0)] or "").strip()
                if not model:
                    continue
                sk = section_key
                if "section_key" in idx:
                    sk = str(row[idx["section_key"]] or section_key).strip()
                result.setdefault(sk, []).append({
                    "model": model,
                    "kw": float(row[idx["kw"]]) if "kw" in idx else None,
                    "q_min": float(row[idx["q_min"]]),
                    "q_max": float(row[idx["q_max"]]),
                    "h_min": float(row[idx["h_min"]]),
                    "h_max": float(row[idx["h_max"]]),
                    "gia": int(row[idx["gia"]]),
                })
            except (KeyError, ValueError, TypeError) as e:
                print(f"Bo qua dong loi: {row[:5]} ({e})")
    return result


def gen_sql(parsed, hang, vat=8, dv="Cụm", markup=1.5):
    total = sum(len(v) for v in parsed.values())
    lines = [
        "-- Tu sinh boi import_bom_catalog.py",
        f"-- Hang: {hang}",
        f"-- Ngay: {datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')}",
        f"-- Tong: {total} model",
        ("-- Quy uoc gia: gia_dau_vao = gia NCC "
         f"| gia_ban_le = gia_dau_vao x {markup:.2f} "
         f"({int(markup*100)}%)"),
        "",
    ]
    for section_key, items in parsed.items():
        if not items:
            continue
        label = SECTION_LABELS.get(section_key, section_key)
        lines.append(f"-- ===== {label} ({len(items)} model) =====")
        for x in items:
            ten = (f"{label} {hang} {x['model']} "
                   f"{x['kw']}kW Q={x['q_min']:.0f}-{x['q_max']:.0f} m³/h "
                   f"H={x['h_min']:.1f}-{x['h_max']:.1f} m")
            ten_sql = ten.replace("'", "''")
            model_sql = x['model'].replace("'", "''")
            hang_sql = hang.replace("'", "''")
            gia_vao = int(x['gia'])
            gia_le = int(round(gia_vao * markup))
            lines.append(
                "INSERT INTO gia_tong_hop "
                "(ten_san_pham, model, nhan_hieu, xuat_xu, don_vi, "
                "gia_dau_vao, gia_ban_le, vat, nhan_cong) VALUES ("
                f"'{ten_sql}', '{model_sql}', '{hang_sql}', "
                f"'Viet Nam', '{dv}', {gia_vao}, {gia_le}, {vat}, 0);")
        lines.append("")
    return "\n".join(lines)


def main():
    ap = argparse.ArgumentParser(
        description="Import bang gia bom chua chay tu PDF/Excel")
    ap.add_argument("input", help="File PDF hoac XLSX")
    ap.add_argument("--hang", required=True, help="Ten hang chu HOA")
    ap.add_argument("--nguon", default="", help="Ten cong ty cung cap")
    ap.add_argument("--ngay",
                    default=datetime.date.today().strftime("%d/%m/%Y"),
                    help="Ngay bao gia")
    ap.add_argument("--merge", action="store_true",
                    help="Merge vao bom_catalog.json")
    ap.add_argument("--vat", type=int, default=8,
                    help="VAT %% (mac dinh 8)")
    ap.add_argument("--dv", default="Cụm",
                    help="Don vi tinh (mac dinh 'Cum')")
    ap.add_argument("--markup", type=float, default=1.5,
                    help="He so nhan gia NCC -> gia ban le (mac dinh 1.5)")
    ap.add_argument("--dry-run", action="store_true",
                    help="Khong ghi file")
    args = ap.parse_args()

    if not os.path.exists(args.input):
        sys.exit(f"Khong tim thay file: {args.input}")

    ext = os.path.splitext(args.input)[1].lower()
    if ext == ".pdf":
        parsed = parse_pdf(args.input)
    elif ext in (".xlsx", ".xls"):
        parsed = parse_xlsx(args.input)
    else:
        sys.exit(f"Dinh dang khong ho tro: {ext}")

    total = sum(len(v) for v in parsed.values())
    print(f"\nDa parse {total} model tu {args.input}")
    for sk, items in parsed.items():
        if items:
            label = SECTION_LABELS.get(sk, sk)
            print(f"  - {label}: {len(items)} model")

    if total == 0:
        sys.exit("Khong trich duoc model nao.")

    print(f"\nGia: gia_dau_vao = gia NCC")
    print(f"     gia_ban_le  = gia_dau_vao x {args.markup:.2f} "
          f"({int(args.markup*100)}%)")

    catalog = {
        "_meta": {
            "hang": args.hang.upper(),
            "nguon": args.nguon,
            "ngay": args.ngay,
        }
    }
    catalog.update(parsed)

    here = os.path.dirname(os.path.abspath(__file__))
    if args.merge:
        json_path = os.path.join(here, "bom_catalog.json")
        existing = {}
        if os.path.exists(json_path):
            try:
                with open(json_path, encoding="utf-8") as f:
                    existing = json.load(f) or {}
            except Exception:
                existing = {}
        if existing.get("_meta", {}).get("hang") == args.hang.upper():
            out = catalog
        else:
            out = existing
            hang_lower = args.hang.lower()
            for sk, items in parsed.items():
                key = f"{hang_lower}_{sk}"
                out[key] = items
            out.setdefault("_meta", {})
            out["_meta"][f"hang_{hang_lower}"] = {
                "nguon": args.nguon, "ngay": args.ngay,
            }
        target_json = json_path
    else:
        target_json = os.path.join(
            here, f"bom_catalog_{args.hang.lower()}.json")
        out = catalog

    sql_text = gen_sql(parsed, args.hang.upper(),
                       vat=args.vat, dv=args.dv, markup=args.markup)
    sql_filename = (f"insert_bom_{args.hang.lower()}_"
                    f"{datetime.datetime.now().strftime('%Y%m%d')}.sql")
    sql_path = os.path.join(here, sql_filename)

    if args.dry_run:
        print(f"\n[DRY-RUN] JSON: {target_json}")
        print(f"[DRY-RUN] SQL:  {sql_path}")
        print("\n----- SQL preview -----")
        print("\n".join(sql_text.splitlines()[:15]))
        return

    with open(target_json, "w", encoding="utf-8") as f:
        json.dump(out, f, indent=2, ensure_ascii=False)
    print(f"\nJSON -> {target_json}")

    with open(sql_path, "w", encoding="utf-8") as f:
        f.write(sql_text)
    print(f"SQL  -> {sql_path}")
    print(f"\nDe nhap vao DB Fsales:")
    print(f"   mysql -u <user> -p <db_name> < {sql_filename}")


if __name__ == "__main__":
    main()
