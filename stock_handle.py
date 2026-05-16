import re
import openpyxl
from PyQt6.QtCore import Qt, QDate
from PyQt6.QtWidgets import QAbstractItemView
from PyQt6.QtGui import QColor
from PyQt6.QtWidgets import QApplication, QMainWindow, QTableWidgetItem, QLineEdit, QPushButton, QStyledItemDelegate
import datetime
from datetime import datetime
import tkinter as tk
from tkinter import filedialog

from autocomplete_utils import setup_autocomplete_for_table_row

import misc
from UI.nhap_xuat_kho import Ui_NhapXuat
from UI.phieu_nhap_kho import Ui_NhapKho
from UI.phieu_xuat_kho_thue import Ui_xuat_kho_thue

import stock_ui_utils

from PyQt6.QtGui import QIntValidator
from ui_theme import apply_ui_v2


class NumericDelegate(QStyledItemDelegate):
    def __init__(self, parent, table_widget):
        super().__init__(parent)
        self.table_widget = table_widget  # Reference to the table for row values

    def createEditor(self, parent, option, index):
        if index.column() == 3:  # Only apply to column 3
            editor = QLineEdit(parent)
            editor.setValidator(QIntValidator(0, 9999999, parent))  # Only allows positive integers
            return editor
        return super().createEditor(parent, option, index)

    def setModelData(self, editor, model, index):
        if index.column() == 3:  # Only validate column 3
            new_value = int(editor.text()) if editor.text().isdigit() else 0
            row = index.row()
            max_value = int(self.table_widget.item(row, 2).text())  # Get value from column 2

            if new_value > max_value:  # Ensure it's <= column 2
                new_value = max_value

            model.setData(index, str(new_value))  # Set validated value
        else:
            super().setModelData(editor, model, index)

def _normalize_key(text: str) -> str:
    if not text:
        return ""
    return text.strip().lower()

class StockHandle(QMainWindow):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.win_stock = QMainWindow()
        self.uic7 = Ui_NhapXuat()
        self.uic7.setupUi(self.win_stock)
        apply_ui_v2(self.win_stock)
        self.win_stock.show()
        self.setWindowTitle(QApplication.translate("Xử lý đơn hàng", "Fsale v2.1.1"))
        self.user = []
        self.user_power = 0
        self.ttkh = None

        self.win_nhap = QMainWindow()
        self.uic8 = Ui_NhapKho()
        self.uic8.setupUi(self.win_nhap)
        apply_ui_v2(self.win_nhap)



    def quan_ly_kho(self):
        self.win_stock = QMainWindow()
        self.uic7 = Ui_NhapXuat()
        self.uic7.setupUi(self.win_stock)
        apply_ui_v2(self.win_stock)
        self.uic7.but_save_phieu_xuat.setHidden(True)
        self.uic7.but_xuat_kho.setHidden(True)
        self.uic7.group_xuatkho.hide()
        self.uic7.but_huy_phieu_xuat.setHidden(True)
        self.uic7.but_tonkho.setHidden(True)
        self.uic7.but_sp_moi.setDisabled(True)

        if not self.user_power:
            try:
                result = misc.sql_one("SELECT power FROM user WHERE full_name = %s", (self.user,))
                if result:
                    self.user_power = int(result[0])
                else:
                    self.user_power = 11
            except Exception as e:
                print("❌ Error getting user power:", e)
                self.user_power = 11

        if self.user_power > 40:
            self.uic7.comboBox_2.addItems(['Phiếu xuất kho', 'Phiếu nhập kho'])
            self.uic7.but_sp_moi.setDisabled(False)
            self.uic7.but_tonkho.setHidden(False)

            px = [str(ele[0]) for ele in misc.sql_all("SELECT id FROM xuat_kho", params=None)]
            self.uic7.combo_sophieu.addItems(px)

            self.uic7.but_duyet.clicked.connect(lambda: StockHandle.duyet_phieu_xuat(self))
            # self.uic7.but_duyet.clicked.connect(lambda: XuatKhoHandle.duyet_phieu_xuat(self.uic7, self.user_power))

            self.uic7.but_xoadong.clicked.connect(lambda: self.uic7.tableWidget.removeRow(self.uic7.tableWidget.rowCount() - 1))
            # self.uic7.but_themdong.clicked.connect(lambda: StockHandle.them_dong(self))
            self.uic7.but_sp_moi.clicked.connect(lambda: StockHandle.san_pham_moi(self))

            self.uic7.comboBox_2.currentTextChanged.connect(lambda: StockHandle.combo_xuat_nhap_change(self))
            self.uic7.combo_sophieu.currentTextChanged.connect(lambda: StockHandle.combo_sophieu_change(self))

        else:
            self.uic7.comboBox_2.addItems(['Phiếu xuất kho'])
            self.uic7.but_duyet.clicked.connect(lambda: StockHandle.phieu_xuat_cua_toi_dau_roi(self))

        if int(self.user_power) > 20:
            self.uic7.but_sp_moi.setDisabled(False)
            self.uic7.but_tonkho.setHidden(False)
        self.uic7.combo_kho.addItems(['Kho Hà Nội', 'Kho HCM'])

        self.uic7.label_user.setText(self.user)

        today = datetime.today()
        self.uic7.dateEdit.setDate(QDate(today.year, today.month, today.day))
        self.win_stock.show()

        StockHandle.xem_ton_kho(self, tenkho=self.uic7.combo_kho.currentText())

        self.uic7.but_tonkho.clicked.connect(lambda: StockHandle.xem_ton_kho(self, tenkho=self.uic7.combo_kho.currentText()))
        self.uic7.but_phieumoi.clicked.connect(lambda: StockHandle.loai_phieu(self))
        self.uic7.but_save_phieu_xuat.clicked.connect(lambda: StockHandle.luu_phieu_xuat(self))

        self.uic7.txt_search.textChanged.connect(lambda: StockHandle.search_kho(self))

    def search_kho(self):
        key = self.uic7.txt_search.toPlainText()
        if '\n' in key:
            key = key.strip()
            self.uic7.txt_search.setText(key)
        else:
            result = []
            kq1 = misc.sql_all("SELECT * FROM ton_kho WHERE model LIKE %s", (f"%{key}%",))
            if kq1:
                result = result + kq1

            kq2 = misc.sql_all("SELECT * FROM ton_kho WHERE ten_san_pham LIKE %s", (f"%{key}%",))
            if kq2:
                result = result + kq2

            if result:
                self.uic7.tableWidget.clear()
                self.uic7.tableWidget.setColumnCount(5)
                self.uic7.tableWidget.setRowCount(len(result))
                self.uic7.tableWidget.setHorizontalHeaderLabels(
                    ['Tên sản phẩm', 'Model', 'tồn kho', 'Giá nhập kho', 'Mã kho'])
                self.uic7.tableWidget.setColumnWidth(0, 350)
                self.uic7.tableWidget.setColumnWidth(1, 100)
                self.uic7.tableWidget.setColumnWidth(2, 90)
                self.uic7.tableWidget.setColumnWidth(3, 90)
                self.uic7.tableWidget.setColumnWidth(3, 100)

                for row in range(len(result)):
                    self.uic7.tableWidget.setItem(row, 0, QTableWidgetItem(result[row][1]))
                    self.uic7.tableWidget.setItem(row, 1, QTableWidgetItem(result[row][2]))
                    self.uic7.tableWidget.setItem(row, 2, QTableWidgetItem(str(result[row][4])))
                    self.uic7.tableWidget.setItem(row, 3, QTableWidgetItem(str(result[row][7])))
                    self.uic7.tableWidget.setItem(row, 4, QTableWidgetItem(str(result[row][8])))

    def phieu_xuat_cua_toi_dau_roi(self):

        kq = misc.sql_all("SELECT * from xuat_kho WHERE kt_duyet = 'F' AND nguoi_lap = %s", (self.user,))

        # Ghi nội dung biến tb - tức là nội dung phiếu xuất - lên màn hình
        self.uic7.tableWidget.clear()
        self.uic7.tableWidget.setColumnCount(5)
        self.uic7.tableWidget.setRowCount(len(kq))
        self.uic7.tableWidget.setHorizontalHeaderLabels(
            ['Số phiếu', 'Nội dung phiếu', 'Người yêu cầu', '   ', 'Ngày lập phiếu'])

        self.uic7.tableWidget.setColumnWidth(0, 60)
        self.uic7.tableWidget.setColumnWidth(1, 380)
        self.uic7.tableWidget.setColumnWidth(2, 100)
        self.uic7.tableWidget.setColumnWidth(3, 100)
        self.uic7.tableWidget.setColumnWidth(4, 100)

        for row in range(len(kq)):
            self.uic7.tableWidget.setItem(row, 0, QTableWidgetItem(str(kq[row][6])))
            self.uic7.tableWidget.setItem(row, 1, QTableWidgetItem(str(kq[row][8])))
            self.uic7.tableWidget.setItem(row, 2, QTableWidgetItem(str(kq[row][3])))

            but1 = QPushButton('Xem đơn')
            # but1.clicked.connect(lambda: StockHandle.ke_toan_duyet(self, kq[self.uic7.tableWidget.currentRow()][0]))

            self.uic7.tableWidget.setCellWidget(row, 3, but1)
            self.uic7.tableWidget.setItem(row, 4, QTableWidgetItem(str(kq[row][1])))

        self.uic7.tableWidget.repaint()

    def combo_sophieu_change(self):
        sophieu = self.uic7.combo_sophieu.currentText()
        if sophieu == '':
            return

        if self.uic7.comboBox_2.currentText() == 'Phiếu xuất kho':
            self.uic7.group_xuatkho.show()

            # Xem lại phieeus xuất theo số phiếu
            kq = misc.sql_all("SELECT * FROM xuat_kho WHERE id = %s", (sophieu,))
            if kq:
                hanghoa = [ele.split('|') for ele in kq[0][2].split('@@')]
                # Ghi nội dung biến hanghoa - tức là nội dung phiếu xuất - lên màn hình
                stock_ui_utils.setup_table_xuat(self.uic7.tableWidget, len(kq))

                for row in range(len(hanghoa)):
                    self.uic7.tableWidget.setItem(row, 0, QTableWidgetItem(str(hanghoa[row][0])))
                    self.uic7.tableWidget.setItem(row, 1, QTableWidgetItem(str(hanghoa[row][1])))
                    self.uic7.tableWidget.setItem(row, 2, QTableWidgetItem(str(hanghoa[row][2])))
                    self.uic7.tableWidget.setItem(row, 3, QTableWidgetItem("{:,}".format(int(hanghoa[row][3]))))
                    self.uic7.tableWidget.setItem(row, 4, QTableWidgetItem(str(hanghoa[row][4])))

                # Hiển thị thông tin phiếu xuất
                txt = 'Phiếu xuất do ' + kq[0][3] + ' lập ngày ' + kq[0][1].strftime("%Y-%m-%d") + '.'
                self.uic7.label_noti.setText(txt)
                self.uic7.label_noti.repaint()

                # Xem lại phieeus xuất theo số phiếu

                kq = misc.sql_all("SELECT * FROM ds_don_hang WHERE so_bg = %s", (kq[0][6],))

                self.uic7.text_so_bg.setText(str(kq[0][0]))
                self.uic7.text_lead.setText(str(kq[0][1]))
                self.uic7.text_gia_tri.setText(str(kq[0][2]))
                self.uic7.text_profit.setText(str(kq[0][5]))
                self.uic7.combo_nguoi_thuc_hien.addItem(kq[0][17])

                # Xem lại phieeus xuất theo số báo giá
                lead_row = misc.sql_one(
                    "SELECT name, sdt, address, company, mst, yc FROM sale_lead WHERE lead_id = %s",
                    (kq[0][1],)
                )
                if lead_row:
                    ten_kh, sdt, dia_chi, ten_cty, mst, yc = lead_row
                    self.uic7.text_noi_dung_xuat.setText((yc or '').strip())
                    self.uic7.text_nguoi_nhan_hang.setText((ten_kh or '').strip())
                    self.uic7.text_sdt.setText((sdt or '').strip())
                    self.uic7.text_dia_chi.setText((dia_chi or '').strip())
                    self.uic7.text_ten_cong_ty.setText((ten_cty or '').strip())
                    self.uic7.text_ma_so_thue.setText((mst or '').strip())

        else:
            try:
                self.uic7.group_xuatkho.hide()
                # Xem lại phieeus xuất theo số phiếu
                kq = misc.sql_all("SELECT * FROM nhap_kho WHERE id = %s", (sophieu,))

                hanghoa = [ele.split('|') for ele in kq[0][2].split('@@')]
                # Hiển thị biến tb - tức là nội dung phiếu nhập lên màn hình
                stock_ui_utils.setup_table_xuat(self.uic7.tableWidget, len(kq))

                for row in range(len(hanghoa)):
                    for col in range(0, 4):
                        if col == 3:
                            self.uic7.tableWidget.setItem(row, col, QTableWidgetItem(
                                "{:,}".format(int(hanghoa[row][col].replace(",", "")))))
                        else:
                            self.uic7.tableWidget.setItem(row, col, QTableWidgetItem(str(hanghoa[row][col])))

                self.uic7.tableWidget.repaint()
            except Exception as e:
                self.uic7.label_noti.setText('Số phiếu mới, chưa có nội dung!')
                self.uic7.label_noti.repaint()
                print(e)

    def combo_xuat_nhap_change(self):
        self.uic7.combo_sophieu.clear()

        if self.uic7.comboBox_2.currentText() == 'Phiếu xuất kho':
            self.uic7.group_xuatkho.show()
            self.uic7.group_xuatkho.setHidden(False)

            # Lấy các số phiếu xuất kho trong quá khứ, xếp theo thứ tự từ lớn đến nhỏ
            px = [str(ele[0]).strip() for ele in misc.sql_all("SELECT id FROM xuat_kho", None)]
        else:
            self.uic7.group_xuatkho.hide()
            # Lấy các số phiếu xuất kho trong quá khứ, xếp theo thứ tự từ lớn đến nhỏ
            px = [str(ele[0]).strip() for ele in misc.sql_all("SELECT id FROM nhap_kho", None)]

        self.uic7.combo_sophieu.addItems(px)

    # def san_pham_moi(self):
    #     self.win_sp = QMainWindow()
    #     self.uic9 = Ui_SanPhamMoi()
    #     self.uic9.setupUi(self.win_sp)
    #     self.win_sp.show()
    #
    #     self.uic9.but_save.clicked.connect(lambda: StockHandle.them_san_pham(self))
    #     self.uic9.but_update.clicked.connect(lambda: StockHandle.update_sp(self))
    #     self.uic9.txt_search.textChanged.connect(lambda: StockHandle.search_sp(self))

    def update_sp(self):
        ttsp = [self.uic9.txt_ten_sp.toPlainText().strip(),
                self.uic9.txt_model.toPlainText().strip(),
                self.uic9.txt_nhan_hieu.toPlainText().strip(),
                self.uic9.txt_xuat_xu.toPlainText().strip(),
                self.uic9.txt_don_vi.toPlainText().strip(),
                str(re.sub(r"\D", "", self.uic9.txt_gia_cap_1.toPlainText())),
                str(re.sub(r"\D", "", self.uic9.txt_gia_cap_2.toPlainText())),
                str(re.sub(r"\D", "", self.uic9.txt_gia_ban_le.toPlainText())),
                str(re.sub(r"\D", "", self.uic9.txt_vat.toPlainText())),
                str(re.sub(r"\D", "", self.uic9.txt_gia_von.toPlainText())),
                str(re.sub(r"\D", "", self.uic9.txt_gia_lap_dat.toPlainText())),
                str(re.sub(r"\D", "", self.uic9.txt_gia_thue_ngay.toPlainText())),
                str(re.sub(r"\D", "", self.uic9.txt_gia_thue_thang.toPlainText())),
                str(re.sub(r"\D", "", self.uic9.txt_gia_thue_nam.toPlainText())),
                self.uic9.txt_mo_ta_sp.toPlainText().strip()]

        try:
            # Kiểm tra xem các thông tin giá và thuế có phải là chữ số ko
            if ttsp[8] in ['8', '10']:
                print('Check thuế xong!')
            else:
                self.uic9.label_noti.setStyleSheet('Color: red')
                self.uic9.label_noti.setText('Thông tin không hợp lệ!')
                self.uic9.label_noti.repaint()
                return

            if any(len(item) for item in ttsp) == 0:
                self.uic9.label_noti.setStyleSheet('Color: red')
                self.uic9.label_noti.setText('Chưa đủ thông tin!')
                self.uic9.label_noti.repaint()
            else:
                check = misc.sql_one("SELECT * FROM gia_tong_hop WHERE model = %s", (ttsp[1],))
                if check:

                    # Ghi thông tin vào file gia_tong_hop
                    misc.sql_commit("UPDATE gia_tong_hop SET ten_san_pham = %s, nhan_hieu = %s, xuat_xu = %s, don_vi = %s, gia_cap_1 = %s,"
                                    "gia_cap_2 = %s, gia_ban_le = %s, vat = %s, gia_dau_vao = %s, nhan_cong = %s, gia_thue_ngay = %s, "
                                    "gia_thue = %s, gia_thue_nam = %s, mo_ta_sp = %s WHERE model = %s",
                                    (ttsp[0], ttsp[2], ttsp[3], ttsp[4], ttsp[5], ttsp[6], ttsp[7], ttsp[8],
                                     ttsp[9], ttsp[10], ttsp[11], ttsp[12], ttsp[13], ttsp[14], ttsp[1],))

                    self.uic9.label_noti.setStyleSheet('Color: blue')
                    self.uic9.label_noti.setText('Đã lưu thông tin sp!')
                    self.uic9.label_noti.repaint()
                else:
                    self.uic9.label_noti.setStyleSheet('Color: red')
                    self.uic9.label_noti.setText('Chưa có model này, không UPDATE được.')
        except Exception as e:
            print(e)

    def search_sp(self):
        txt = self.uic9.txt_search.toPlainText()
        if '\n' not in txt:
            return
        else:
            sp = txt.strip()
            self.uic9.txt_search.setText(sp)
            kq = misc.sql_one("SELECT * FROM gia_tong_hop WHERE model = %s", (sp,))
            if kq:
                self.uic9.txt_ten_sp.setText(kq[1])
                self.uic9.txt_model.setText(kq[2])
                self.uic9.txt_nhan_hieu.setText(kq[3])
                self.uic9.txt_xuat_xu.setText(kq[4])
                self.uic9.txt_don_vi.setText(kq[5])
                self.uic9.txt_gia_cap_1.setText(str(kq[6]))
                self.uic9.txt_gia_cap_2.setText(str(kq[7]))
                self.uic9.txt_gia_ban_le.setText(str(kq[8]))
                self.uic9.txt_vat.setText(str(kq[9]))
                self.uic9.txt_gia_von.setText(str(kq[10]))
                self.uic9.txt_gia_lap_dat.setText(str(kq[11]))
                self.uic9.txt_gia_thue_ngay.setText(str(kq[12]))
                self.uic9.txt_gia_thue_thang.setText(str(kq[13]))
                self.uic9.txt_gia_thue_nam.setText(str(kq[14]))
                self.uic9.txt_mo_ta_sp.setText(kq[15])

                self.uic9.label_noti.clear()

            else:
                self.uic9.label_noti.setStyleSheet('color: red')
                self.uic9.label_noti.setText('Không tìm thấy model này')

    def them_san_pham(self):
        ttsp = [self.uic9.txt_ten_sp.toPlainText().strip(),
                self.uic9.txt_model.toPlainText().strip(),
                self.uic9.txt_nhan_hieu.toPlainText().strip(),
                self.uic9.txt_xuat_xu.toPlainText().strip(),
                self.uic9.txt_don_vi.toPlainText().strip(),
                str(re.sub(r"\D", "", self.uic9.txt_gia_cap_1.toPlainText())),
                str(re.sub(r"\D", "", self.uic9.txt_gia_cap_2.toPlainText())),
                str(re.sub(r"\D", "", self.uic9.txt_gia_ban_le.toPlainText())),
                str(re.sub(r"\D", "", self.uic9.txt_vat.toPlainText())),
                str(re.sub(r"\D", "", self.uic9.txt_gia_von.toPlainText())),
                str(re.sub(r"\D", "", self.uic9.txt_gia_lap_dat.toPlainText())),
                str(re.sub(r"\D", "", self.uic9.txt_gia_thue_ngay.toPlainText())),
                str(re.sub(r"\D", "", self.uic9.txt_gia_thue_thang.toPlainText())),
                str(re.sub(r"\D", "", self.uic9.txt_gia_thue_nam.toPlainText())),
                self.uic9.txt_mo_ta_sp.toPlainText().strip()]

        try:
            # Kiểm tra xem các thông tin giá và thuế có phải là chữ số ko
            if ttsp[8] in ['8', '10']:
                # Nếu đúng rồi thì kiểm tra xem model có trong kho chưa
                kq = misc.sql_all("SELECT ten_san_pham, model from gia_tong_hop", None)

                tex1 = ''
                if ttsp[0].lower() in [item[0].lower() for item in kq]:
                    tex1 = 'Có sản phẩm trùng tên.'
                if ttsp[1].lower() in [item[1].lower() for item in kq]:
                    tex2 = 'Trùng model với sản phẩm cũ' + tex1
                    self.uic9.label_noti.setStyleSheet('Color: red')
                    self.uic9.label_noti.setText(tex2)
                    self.uic9.label_noti.repaint()
                    return

            else:
                self.uic9.label_noti.setStyleSheet('Color: red')
                self.uic9.label_noti.setText('Thông tin không hợp lệ!')
                self.uic9.label_noti.repaint()
                return

            if any(len(item) for item in ttsp) == 0:
                self.uic9.label_noti.setStyleSheet('Color: red')
                self.uic9.label_noti.setText('Chưa đủ thông tin!')
                self.uic9.label_noti.repaint()
            else:
                # Ghi thông tin vào file gia_tong_hop
                misc.sql_commit("INSERT INTO gia_tong_hop (ten_san_pham, model, nhan_hieu, xuat_xu, don_vi, gia_cap_1,"
                                "gia_cap_2, gia_ban_le, vat, gia_dau_vao, nhan_cong, gia_thue_ngay, gia_thue, gia_thue_nam, mo_ta_sp) "
                                "VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)",
                                (ttsp[0], ttsp[1], ttsp[2], ttsp[3], ttsp[4], ttsp[5], ttsp[6], ttsp[7], ttsp[8], ttsp[9], ttsp[10], ttsp[11], ttsp[12], ttsp[13], ttsp[14]))

                self.uic9.label_noti.setStyleSheet('Color: blue')
                self.uic9.label_noti.setText('Đã lưu thông tin sp!')
                self.uic9.label_noti.repaint()
        except Exception as e:
            print(e)

    def load_gia_nhap(self):
        # Nhận số liệu từ màn hình và ghi vào 1 list tên là tb
        warehouse_name = self.uic8.combo_kho.currentText()
        rows = self.uic8.tableWidget.rowCount()
        cols = self.uic8.tableWidget.columnCount()
        tb = []
        for row in range(rows):
            if self.uic8.tableWidget.item(row, 1) is None and self.uic8.tableWidget.item(row, 2) is None:
                pass
            else:
                row_content = []
                for col in range(cols):
                    item = self.uic8.tableWidget.item(row, col)

                    if item is not None:
                        row_content.append(item.text().strip())
                    else:
                        row_content.append('None')
                row_content.append(warehouse_name)
                tb.append(row_content)

        tb = [item for item in tb if item[1] != '' and item[2] != 'None']

        for item in tb:
            item[3] = misc.sql_one("SELECT gia_dau_vao FROM gia_tong_hop WHERE model = %s", (item[1],))[0]

        # Hiển thị biến tb - tức là nội dung phiếu nhập lên màn hình
        self.uic8.tableWidget.clear()
        self.uic8.tableWidget.setColumnCount(4)
        self.uic8.tableWidget.setRowCount(len(tb))
        self.uic8.tableWidget.setHorizontalHeaderLabels(
            ['Tên sản phẩm', 'Model', 'Số lượng', 'Đơn giá'])

        self.uic8.tableWidget.setColumnWidth(0, 380)
        self.uic8.tableWidget.setColumnWidth(1, 100)
        self.uic8.tableWidget.setColumnWidth(2, 80)
        self.uic8.tableWidget.setColumnWidth(3, 100)

        for row in range(len(tb)):
            for col in range(0, 4):
                if col == 3:
                    self.uic8.tableWidget.setItem(row, col, QTableWidgetItem("{:,}".format(int(tb[row][col]))))
                else:
                    self.uic8.tableWidget.setItem(row, col, QTableWidgetItem(str(tb[row][col])))

        self.uic8.tableWidget.repaint()

    def duyet_phieu_xuat(self):
        self.uic7.but_xuat_kho.setHidden(True)
        self.uic7.but_huy_phieu_xuat.setHidden(True)
        self.uic7.but_save_phieu_xuat.setHidden(True)
        self.uic7.group_xuatkho.hide()


        kq = misc.sql_all("SELECT * FROM xuat_kho WHERE kt_duyet = 'F' ORDER BY id ASC", None)

        # Ghi nội dung biến tb - tức là nội dung phiếu xuất - lên màn hình
        stock_ui_utils.setup_table_xuat(self.uic7.tableWidget, len(kq))

        for row in range(len(kq)):
            self.uic7.tableWidget.setItem(row, 0, QTableWidgetItem(str(kq[row][6])))
            self.uic7.tableWidget.setItem(row, 1, QTableWidgetItem(str(kq[row][8])))
            self.uic7.tableWidget.setItem(row, 2, QTableWidgetItem(str(kq[row][3])))

            if row == 0:
                but1 = QPushButton('Xem đơn')
                but1.clicked.connect(lambda: StockHandle.ke_toan_duyet(self, kq[self.uic7.tableWidget.currentRow()][0]))

                self.uic7.tableWidget.setCellWidget(row, 3, but1)
                self.uic7.tableWidget.setItem(row, 4, QTableWidgetItem(str(kq[row][1])))

    def show_phieu_xuat(self, sophieu):
        kq = misc.sql_one("SELECT * from xuat_kho WHERE id = %s", (sophieu,))

        hanghoa = [ele.split('|') for ele in kq[2].split('@@')]

        # Ghi nội dung biến tb - tức là nội dung phiếu xuất - lên màn hình
        stock_ui_utils.setup_table_xuat(self.uic7.tableWidget, len(kq))

        for row in range(len(hanghoa)):
            self.uic7.tableWidget.setItem(row, 0, QTableWidgetItem(str(hanghoa[row][0])))
            self.uic7.tableWidget.setItem(row, 1, QTableWidgetItem(str(hanghoa[row][1])))
            self.uic7.tableWidget.setItem(row, 2, QTableWidgetItem(str(hanghoa[row][2])))
            self.uic7.tableWidget.setItem(row, 3, QTableWidgetItem("{:,}".format(int(hanghoa[row][3]))))
            self.uic7.tableWidget.setItem(row, 4, QTableWidgetItem(str(hanghoa[row][4])))

        self.uic7.tableWidget.repaint()

        self.uic7.text_noi_dung_xuat.setText(kq[8])
        self.uic7.text_lead.setText(str(kq[7]))
        self.uic7.text_so_bg.setText(str(kq[6]))
        lead_id = kq[7]
        lead_row = misc.sql_one("SELECT name, sdt, address, company, mst FROM sale_lead WHERE lead_id = %s",(lead_id,))
        if lead_row:
            ten_kh, sdt, dia_chi, ten_cty, mst = lead_row
            self.uic7.text_nguoi_nhan_hang.setText((ten_kh or '').strip())
            self.uic7.text_sdt.setText((sdt or '').strip())
            self.uic7.text_dia_chi.setText((dia_chi or '').strip())
            self.uic7.text_ten_cong_ty.setText((ten_cty or '').strip())
            self.uic7.text_ma_so_thue.setText((mst or '').strip())

    def ke_toan_duyet(self, sophieu):
        cong_no = 0

        if int(self.user_power) < 40:
            return

        self.uic7.but_huy_phieu_xuat.setHidden(False)
        self.uic7.but_xuat_kho.setHidden(False)
        self.uic7.group_xuatkho.setHidden(False)

        kq = misc.sql_one("SELECT * from xuat_kho WHERE id = %s", (sophieu,))
        hanghoa = [ele.split('|') for ele in kq[2].split('@@')]

        StockHandle.show_phieu_xuat(self, sophieu)

        kq1 = misc.sql_one("SELECT * from ds_don_hang WHERE so_bg = %s", (kq[6],))
        if kq1:
            self.uic7.text_profit.setText("{:,}".format(kq1[5]))
            self.uic7.text_gia_tri.setText("{:,}".format(kq1[2] + kq1[3]))
            self.uic7.text_cong_no.setText("{:,}".format(kq1[18]))

            if kq1[18] <= 0:
                self.uic7.checkBox.setDisabled(True)
            else:
                self.uic7.checkBox.setDisabled(False)

            self.uic7.combo_nguoi_thuc_hien.addItem(kq1[17])

        else:
            kq1 = misc.sql_one("SELECT * from ds_don_thue WHERE so_bg = %s", (kq[6],))
            self.uic7.text_profit.setText("{:,}".format(kq1[7]))
            self.uic7.text_gia_tri.setText("{:,}".format(kq1[2]))
            cong_no = int(kq1[2]) + int(kq1[3]) + int(kq1[5])
            self.uic7.text_cong_no.setText("{:,}".format(cong_no))

            if cong_no <= 0:
                self.uic7.checkBox.setDisabled(True)
            else:
                self.uic7.checkBox.setDisabled(False)

            self.uic7.combo_nguoi_thuc_hien.addItem(kq1[16])

        if cong_no > 0 and not self.uic7.checkBox.isChecked():
            self.uic7.but_xuat_kho.setDisabled(True)
            self.uic7.checkBox.stateChanged.connect(StockHandle.check_dong_y(self, cong_no))

        self.uic7.but_xuat_kho.clicked.connect(lambda: StockHandle.xuat_file(self, sophieu))

        self.uic7.but_huy_phieu_xuat.clicked.connect(lambda: StockHandle.huy_phieu_xuat(self, sophieu, hanghoa))

    def check_dong_y(self, cong_no):
        if cong_no > 0 and not self.uic7.checkBox.isChecked():
            self.uic7.but_xuat_kho.setDisabled(True)

        if cong_no == 0:
            self.uic7.but_xuat_kho.setDisabled(False)

        if cong_no > 0 and self.uic7.checkBox.isChecked():
            self.uic7.but_xuat_kho.setDisabled(False)

    def huy_phieu_xuat(self, sophieu, hanghoa):

        # Xóa số phiếu trong db
        misc.sql_commit("DELETE from xuat_kho WHERE id = %s", (sophieu,))

        # # Sửa lại nội dung file xuất - nhập - tồn
        for item in hanghoa:
            if item[1] not in ['APP', 'NhanCong']:
                kq = misc.sql_one(f"SELECT * FROM ton_kho WHERE model = %s AND ma_kho = %s", (item[1], item[4],))
                misc.sql_commit("UPDATE ton_kho SET ton = %s, xuat = %s WHERE ma_kho = %s AND model = %s", (int(kq[4]) + int(item[2]), int(kq[5]) - int(item[2]), item[4], item[1],))

        self.uic7.label_noti.setStyleSheet('color: red')
        self.uic7.label_noti.setText('Đã XÓA phiếu xuất số ' + str(sophieu))
        self.uic7.label_noti.repaint()
        txt = str(self.user) + ' đã từ chối xuất kho phiếu xuất số ' + str(sophieu)
        misc.send_to_telegram(txt)
        StockHandle.duyet_phieu_xuat(self)
        # XuatKhoHandle.duyet_phieu_xuat(self.uic7, self.user_power)

    def loai_phieu(self):
        self.uic7.but_xuat_kho.setHidden(True)
        if self.uic7.comboBox_2.currentText() == 'Phiếu nhập kho':
            self.uic7.but_save_phieu_xuat.setHidden(True)
            StockHandle.tao_phieu_nhap(self)
        else:
            self.uic7.but_save_phieu_xuat.setHidden(False)
            self.uic7.group_xuatkho.show()
            StockHandle.tao_phieu_xuat(self)

    def tao_phieu_xuat_tu_don_hang(self, lead_id, so_bg):
        # ttkh = ['tên công ty',ten_khách, số đt, lead_id, noi dung yeu cau, 'mã số thuế']
        ttkh = misc.sql_one("SELECT company, name, sdt, yc, mst, address FROM sale_lead WHERE lead_id = %s", (lead_id,))
        ten_cty, ten_kh, sdt, yc, mst, dia_chi = ttkh

        StockHandle.quan_ly_kho(self)

        self.uic7.but_xuat_kho.setHidden(True)
        self.uic7.but_save_phieu_xuat.setHidden(False)
        self.uic7.group_xuatkho.setHidden(False)
        self.uic7.but_huy_phieu_xuat.setHidden(True)
        self.uic7.but_themdong.setDisabled(True)
        self.uic7.but_xoadong.setDisabled(True)

        self.uic7.label.setText('Phiếu xuất kho')
        # self.uic7.label.repaint()
        # Điền thông tin khách hàng

        self.uic7.text_nguoi_nhan_hang.setText((ten_kh or '').strip())
        self.uic7.text_sdt.setText((sdt or '').strip())
        self.uic7.text_dia_chi.setText((dia_chi or '').strip())
        self.uic7.text_ten_cong_ty.setText((ten_cty or '').strip())
        self.uic7.text_ma_so_thue.setText((mst or '').strip())

        dien_giai = misc.sql_one("SELECT tieu_de FROM ds_bao_gia WHERE so_bg = %s", (so_bg,))[0]
        self.uic7.text_noi_dung_xuat.setText(dien_giai)
        self.uic7.text_so_bg.setText(str(so_bg))
        self.uic7.text_lead.setText(str(lead_id))

        # Khởi tạo header
        self.uic7.combo_kho.addItems(['Kho Hà Nội', 'Kho HCM'])
        self.uic7.combo_kho.setCurrentText('Kho Hà Nội')
        self.uic7.comboBox_2.setCurrentText('Phiếu xuất kho')
        self.uic7.label_user.setText(self.user)
        txt = 'Báo giá số ' + str(so_bg) + ': ' + ttkh[3]
        self.uic7.label_noti.setStyleSheet('color: blue')
        self.uic7.label_noti.setText(txt)

        kq = misc.sql_all("SELECT full_name FROM user WHERE power > 0 and power < 50", None)
        # Lấy danh sách user điền vào ô người giao nhận
        kq1 = list(kq)
        self.uic7.combo_nguoi_thuc_hien.addItem('Người setup - giao nhận')
        self.uic7.combo_nguoi_thuc_hien.addItems(item[0] for item in kq1)

        kq = misc.sql_one('SELECT * from xuat_kho WHERE so_bg = %s', (so_bg,))

        if kq:
            # Lấy số phiếu xuất
            # so_phieu_xuat = kq[0]
            self.uic7.combo_sophieu.clear()
            self.uic7.combo_sophieu.addItems([str(kq[0])])

            # Lấy danh sách hàng hóa
            hh = misc.sql_one("SELECT noi_dung FROM ds_bao_gia WHERE so_bg = %s", (so_bg,))[0]
            ds = [item.split('|') for item in hh.split('@')]
            hanghoa = [ele for ele in ds if ele != ['']]

        else:
            # Lấy số phiếu xuất kho mới
            kq = misc.sql_one("SELECT * FROM xuat_kho WHERE id = (SELECT MAX(id) FROM xuat_kho)")
            try:
                self.uic7.combo_sophieu.clear()
                self.uic7.combo_sophieu.addItems([str(kq[0] + 1)])
            except Exception as e:
                print(e)

            # Lấy danh mục hàng hóa và hiển thị
            kq = misc.sql_one("SELECT noi_dung FROM ds_bao_gia WHERE so_bg = %s", (so_bg,))
            ds = [item.split('|') for item in kq[0].split('@')]
            hanghoa = [ele for ele in ds if ele != ['']]

            self.uic7.tableWidget.setRowCount(0)
            StockHandle.khoi_tao(self, uic=self.uic7)

        # Ghi nội dung biến hanghoa - tức là nội dung phiếu xuất - lên màn hình
        stock_ui_utils.setup_table_xuat(self.uic7.tableWidget, len(hanghoa))

        for row in range(len(hanghoa)):
            if hanghoa[row] == ['']:
                pass
            else:
                self.uic7.tableWidget.setItem(row, 0, QTableWidgetItem(str(hanghoa[row][0])))
                self.uic7.tableWidget.setItem(row, 1, QTableWidgetItem(str(hanghoa[row][1])))
                self.uic7.tableWidget.setItem(row, 2, QTableWidgetItem(str(hanghoa[row][4])))
                self.uic7.tableWidget.setItem(row, 3, QTableWidgetItem("{:,}".format(int(hanghoa[row][5].replace(",", "")))))

        kq = misc.sql_one(f'SELECT * from ds_don_hang WHERE so_bg = %s', (so_bg,))
        nguoi_setup = kq[17]
        if nguoi_setup in kq1:
            self.uic7.combo_nguoi_thuc_hien.setCurrentText(nguoi_setup)

        tien_hang = kq[2]
        vat = kq[3]
        self.uic7.text_gia_tri.setText("{:,}".format(tien_hang + vat))

        da_thanh_toan = kq[4]
        con_no = tien_hang + vat - da_thanh_toan
        self.uic7.text_cong_no.setText("{:,}".format(con_no))

    def tao_phieu_xuat(self):

        self.uic7.but_xuat_kho.setHidden(True)
        self.uic7.but_save_phieu_xuat.setHidden(True)  # Mở lại khi code đêến phần tạo phiếu xuaats kho ở đây
        self.uic7.group_xuatkho.show()

        self.uic7.but_themdong.setDisabled(True)
        self.uic7.but_xoadong.setDisabled(True)
        self.uic7.label.setText('Phiếu xuất kho')
        self.uic7.label.repaint()
        self.uic7.tableWidget.clear()

        self.uic7.label_noti.setStyleSheet("color: red")
        self.uic7.label_noti.setText('Chỉ tạo được phiếu xuất từ đơn hàng - hãy tạo đơn hàng trước!')
        self.uic7.label_noti.repaint()

    def tao_phieu_nhap(self):
        self.win_nhap = QMainWindow()
        self.uic8 = Ui_NhapKho()
        self.uic8.setupUi(self.win_nhap)
        apply_ui_v2(self.win_nhap)
        self.win_nhap.show()

        self.uic8.dateEdit.setDate(QDate.currentDate())
        self.uic8.combo_so_px.addItems(['0'])
        self.uic8.label_thong_tin.hide()
        self.uic8.check_bao_hanh.hide()
        self.uic8.check_tra_lai.hide()
        self.uic8.combo_so_px.currentTextChanged.connect(lambda: StockHandle.on_so_px_change(self))
        self.uic8.but_save_NHAP_tu_p_xuat.hide()
        self.uic8.but_save_phieu_NHAP.show()
        self.uic8.combo_kho.addItems(["Kho Hà Nội", "Kho HCM"])

        StockHandle.khoi_tao(self, self.uic8)

        kq = misc.sql_one("SELECT * FROM nhap_kho WHERE id = (SELECT MAX(id) FROM nhap_kho)")
        self.uic8.combo_sophieu.clear()
        try:
            self.uic8.combo_sophieu.addItems([str(kq[0] + 1)])
        except:
            self.uic8.combo_sophieu.addItems(['1'])

        self.uic8.but_save_NHAP_tu_p_xuat.clicked.connect(lambda: StockHandle.luu_phieu_nhap_lai_tu_phieu_xuat(self))
        self.uic8.but_save_phieu_NHAP.clicked.connect(lambda: StockHandle.luu_phieu_nhap(self))
        self.uic8.but_load_gia.clicked.connect(lambda: StockHandle.load_gia_nhap(self))
        self.uic8.checkBox.stateChanged.connect(lambda: StockHandle.on_check_phieu_xuat_changed(self))

    def on_so_px_change(self):

        user = misc.sql_one("SELECT nguoi_lap, so_bg, ngay_thang FROM xuat_kho WHERE id = %s", (self.uic8.combo_so_px.currentText(),))
        if not user:
            self.uic8.label_thong_tin.setText('Không tìm thấy phiếu xuất tương ứng.')
            return

        don_hang = misc.sql_one("SELECT * FROM ds_don_hang WHERE so_bg = %s", (user[1],))
        if not don_hang:
            self.uic8.label_thong_tin.setText('Không tìm thấy đơn hàng tương ứng với phiếu xuất.')
            return

        ngay_xuat = user[2]
        if hasattr(ngay_xuat, 'strftime'):
            ngay_xuat = ngay_xuat.strftime('%d/%m/%Y %H:%M:%S')
        else:
            ngay_xuat = str(ngay_xuat)

        self.uic8.label_thong_tin.setText(
            'của ' + str(user[0]) + ' theo báo giá số ' + str(don_hang[0]) + ' đã xuất hàng ngày ' + ngay_xuat
        )

        ds = [ele.split('|') for ele in don_hang[10].split('@@')]

        self.uic8.tableWidget.clear()
        self.uic8.tableWidget.setColumnCount(5)
        self.uic8.tableWidget.setRowCount(len(ds))
        self.uic8.tableWidget.setHorizontalHeaderLabels(
            ['Tên sản phẩm', 'Model', 'Số lượng', 'Giá theo đơn xuất', 'Số lượng nhập lại'])

        self.uic8.tableWidget.setColumnWidth(0, 380)
        self.uic8.tableWidget.setColumnWidth(1, 100)
        self.uic8.tableWidget.setColumnWidth(2, 80)
        self.uic8.tableWidget.setColumnWidth(3, 100)
        self.uic8.tableWidget.setColumnWidth(4, 100)

        for row in range(len(ds)):
            for col in [0, 1, 2, 3]:  # Disable editing for these columns
                item = QTableWidgetItem(str(ds[row][col]))
                item.setFlags(Qt.ItemFlag.ItemIsSelectable | Qt.ItemFlag.ItemIsEnabled)  # Read-only
                self.uic8.tableWidget.setItem(row, col, item)

            # Column 3 remains editable
            item = QTableWidgetItem('0')
            item.setFlags(Qt.ItemFlag.ItemIsSelectable | Qt.ItemFlag.ItemIsEnabled | Qt.ItemFlag.ItemIsEditable)
            self.uic8.tableWidget.setItem(row, 4, item)

            delegate = NumericDelegate(self.uic8.tableWidget, self.uic8.tableWidget)
            self.uic8.tableWidget.setItemDelegateForColumn(4, delegate)

    def on_check_phieu_xuat_changed(self):
        if self.uic8.checkBox.isChecked():
            self.uic8.but_save_NHAP_tu_p_xuat.show()
            self.uic8.but_save_phieu_NHAP.hide()
            self.uic8.combo_so_px.setEnabled(True)

            ds = sorted([ele[0] for ele in misc.sql_all("SELECT id FROM xuat_kho")], reverse=True)
            ds = [str(ele) for ele in ds]
            self.uic8.combo_so_px.addItems(ds)
            index = self.uic8.combo_so_px.findText("0")  # Find index of the item
            if index != -1:
                self.uic8.combo_so_px.removeItem(index)  # Remove item by index
            self.uic8.label_thong_tin.show()
            self.uic8.check_bao_hanh.show()
            self.uic8.check_tra_lai.show()
            self.uic8.but_load_gia.hide()

        else:
            self.uic8.but_save_NHAP_tu_p_xuat.hide()
            self.uic8.but_save_phieu_NHAP.show()
            self.uic8.combo_so_px.setEnabled(False)
            self.uic8.label_thong_tin.clear()
            StockHandle.khoi_tao(self, uic=self.uic8)
            self.uic8.but_load_gia.show()
            self.uic8.check_bao_hanh.hide()
            self.uic8.check_tra_lai.hide()
            self.uic8.label_thong_tin.hide()

    def luu_phieu_nhap_lai_tu_phieu_xuat(self):
        print('Lưu phiếu nhập lại từ phiếu xuất')
        # Load phiếu nhập
        data_list = []
        row_count = self.uic8.tableWidget.rowCount()

        for row in range(row_count):
            row_data = []
            for col in [0, 1, 3, 4]:
                item = self.uic8.tableWidget.item(row, col)
                row_data.append(item.text() if item else "")  # Get text, or empty string if None
            data_list.append(row_data)

        # Ghi nhận số liệu trên màn hình vào một biến gọi là hanghoa
        hanghoa = []
        for row in range(row_count):
            row_content = []
            for col in [0, 1, 3, 4]:
                item = self.uic8.tableWidget.item(row, col)
                row_content.append(item.text().strip())
            row_content.append(row_content[3])
            row_content[3] = row_content[2]
            row_content[2] = row_content[4]
            row_content[4] = self.uic8.combo_kho.currentText()

            if self.uic8.check_bao_hanh.isChecked():
                row_content[2] = 0
            hanghoa.append(row_content)
        # Đã ghi xong hàng hóa
        # Tạo phiếu nhập và ghi nội dung vào data base
        warehouse_name = self.uic8.combo_kho.currentText()
        so_pn = self.uic8.combo_sophieu.currentText()
        ngaythang = self.uic8.dateEdit.text()
        current_time = datetime.now().strftime("%H:%M")
        ghi_chu = self.uic8.text_nguyen_nhan.toPlainText()
        if len(ghi_chu) < 7 and not self.uic8.check_bao_hanh.isChecked():
            self.uic8.label_noti.setStyleSheet('color: red')
            self.uic8.label_noti.setText('Cần phải ghi rõ lý do nhập lại hàng')
            return
        if self.uic8.check_bao_hanh.isChecked():
            ghi_chu = 'Nhập lại để bảo hành ' + ghi_chu

        so_px_ref = self.uic8.combo_so_px.currentText().strip()

        # PHASE 3: validate số lượng nhập lại không vượt quá số lượng đã xuất theo từng model
        # 1) Tổng số lượng đã xuất theo model từ phiếu xuất tham chiếu
        xuat_by_model = {}
        for item in hanghoa:
            try:
                model = str(item[1]).strip()
                sl_xuat = int(str(item[3]).replace(',', '').strip() or '0')
                xuat_by_model[model] = sl_xuat
            except Exception:
                pass

        # 2) Tổng số lượng yêu cầu nhập lại lần này theo model
        nhap_now_by_model = {}
        for item in hanghoa:
            try:
                model = str(item[1]).strip()
                sl_nhap = int(str(item[2]).replace(',', '').strip() or '0')
                if sl_nhap < 0:
                    self.uic8.label_noti.setStyleSheet('color: red')
                    self.uic8.label_noti.setText(f'Số lượng nhập lại không hợp lệ (model: {model}).')
                    return
                nhap_now_by_model[model] = nhap_now_by_model.get(model, 0) + sl_nhap
            except Exception:
                self.uic8.label_noti.setStyleSheet('color: red')
                self.uic8.label_noti.setText('Dữ liệu số lượng nhập lại không hợp lệ.')
                return

        # 3) Tổng số lượng đã nhập lại trước đó của cùng phiếu xuất (nếu có)
        nhap_prev_by_model = {}
        if so_px_ref:
            prev_rows = misc.sql_all(
                "SELECT noi_dung FROM nhap_kho WHERE ghi_chu LIKE %s",
                (f"%TRA_LAI_PX:{so_px_ref}%",)
            )
            for r in prev_rows or []:
                nd = (r[0] or '').split('@@')
                for line in nd:
                    cols = line.split('|')
                    if len(cols) >= 3:
                        model = str(cols[1]).strip()
                        try:
                            sl = int(str(cols[2]).replace(',', '').strip() or '0')
                        except Exception:
                            sl = 0
                        nhap_prev_by_model[model] = nhap_prev_by_model.get(model, 0) + sl

        # 4) So sánh giới hạn
        for model, sl_now in nhap_now_by_model.items():
            sl_xuat = int(xuat_by_model.get(model, 0))
            sl_prev = int(nhap_prev_by_model.get(model, 0))
            if sl_prev + sl_now > sl_xuat:
                self.uic8.label_noti.setStyleSheet('color: red')
                self.uic8.label_noti.setText(
                    f'Vượt số lượng xuất cho model {model}: đã nhập lại {sl_prev}, đang nhập {sl_now}, tối đa {sl_xuat}.'
                )
                return

        if so_px_ref:
            ghi_chu = f"TRA_LAI_PX:{so_px_ref} | " + ghi_chu

        ds = []
        for item in hanghoa:
            ds.append('|'.join(item))
        tb1 = "@@".join(ds)

        sql_query = """
                INSERT INTO nhap_kho (id, ngay_thang, noi_dung, nguoi_lap, ma_kho, gio_nhap, ghi_chu)
                VALUES (%s, %s, %s, %s, %s, %s, %s)
                """
        params = (so_pn, ngaythang, tb1, self.user, warehouse_name, current_time, ghi_chu,)
        misc.sql_commit(sql_query, params)

        self.uic8.label_noti.setStyleSheet('color: blue')
        self.uic8.label_noti.setText('Đã thêm mới phiếu nhập kho lại từ đơn hàng đã xuất ')

        # Cộng số lượng nhập kho mới vào file ton_kho trên DB

        for item in hanghoa:
            # Nếu chưa có model này trong kho này
            kq = misc.sql_one(f"SELECT * from ton_kho WHERE model = %s AND ma_kho = %s", (item[1], warehouse_name))
            if not kq:
                misc.sql_commit(
                    "INSERT INTO ton_kho (ten_san_pham, model, nhap, gia_nhap, ma_kho, ton) "
                    "VALUES (%s, %s, %s, %s, %s, %s)",
                    (item[0], item[1], item[2], item[3].replace(",", ""), warehouse_name, item[2])
                )

            else:
                sl_ton = int(kq[4])
                sl_da_nhap_kho = int(kq[6])
                gia_tb = int(kq[7])
                sl_nhap = int(item[2])
                gia_nhap_moi = int(item[3].replace(",", ""))
                gia_tb_moi = (sl_da_nhap_kho * gia_tb + sl_nhap * gia_nhap_moi) // (sl_da_nhap_kho + sl_nhap)

                misc.sql_commit(
                    f"UPDATE ton_kho SET nhap = %s, ton = %s, gia_nhap = %s WHERE model = %s AND ma_kho = %s",
                    (sl_nhap + sl_da_nhap_kho, sl_ton + sl_nhap, gia_tb_moi, item[1], warehouse_name))

    def luu_phieu_nhap(self):
        print('Đang lưu phiếu nhập hàng vào kho')
        warehouse_name = self.uic8.combo_kho.currentText()
        today = QDate.currentDate()
        self.uic8.dateEdit.setDate(today)

        # Nhận số liệu từ màn hình và ghi vào 1 list tên là tb
        rows = self.uic8.tableWidget.rowCount()
        cols = self.uic8.tableWidget.columnCount()
        tb = []
        for row in range(rows):
            if self.uic8.tableWidget.item(row, 1) is None or self.uic8.tableWidget.item(row, 2) is None or self.uic8.tableWidget.item(row, 3) is None:
                pass
            else:
                row_content = []
                for col in range(cols):
                    item = self.uic8.tableWidget.item(row, col)

                    if item is not None:
                        if item.text().strip() == '':   # Nếu chưa có thông tin về sản phẩm, autocompleter sẽ trả về ''
                            self.uic8.label_noti.setStyleSheet('color: red')
                            self.uic8.label_noti.setText('Chưa có sản phẩm này trong dữ liệu. Cần bổ sung thông tin về SP mới trước khi nhập kho.')
                            self.uic8.label_noti.repaint()
                            return

                        row_content.append(item.text().strip())
                    else:
                        row_content.append('None')
                row_content.append(warehouse_name)
                tb.append(row_content)

        # Hiển thị biến tb - tức là nội dung phiếu nhập lên màn hình
        self.uic8.tableWidget.clear()
        self.uic8.tableWidget.setColumnCount(4)
        self.uic8.tableWidget.setRowCount(len(tb))
        self.uic8.tableWidget.setHorizontalHeaderLabels(
            ['Tên sản phẩm', 'Model', 'Số lượng', 'Đơn giá'])

        self.uic8.tableWidget.setColumnWidth(0, 380)
        self.uic8.tableWidget.setColumnWidth(1, 100)
        self.uic8.tableWidget.setColumnWidth(2, 80)
        self.uic8.tableWidget.setColumnWidth(3, 100)

        for row in range(len(tb)):
            for col in range(0, 4):
                if col == 3:
                    self.uic8.tableWidget.setItem(row, col, QTableWidgetItem("{:,}".format(int(tb[row][col].replace(",", "")))))
                else:
                    self.uic8.tableWidget.setItem(row, col, QTableWidgetItem(str(tb[row][col])))

        self.uic8.tableWidget.repaint()

        # Ghi nội dung phiếu nhập kho vào file nhap_kho trên DB
        sophieu = self.uic8.combo_sophieu.currentText()

        ngaythang = self.uic8.dateEdit.text()
        # makho = self.uic8.combo_kho.currentText()
        ds = []

        for item in tb:
            ds.append('|'.join(item))
        tb1 = "@@".join(ds)

        kq = misc.sql_one("SELECT * FROM nhap_kho WHERE id = %s", (sophieu,))

        current_time = datetime.now().strftime("%H:%M")
        if kq:
            noi_dung_phieu_cu = [ele.split('|') for ele in kq[2].split('@@')]
            # Nếu là sửa file cũ - trừ đi số lượng cũ trong file tồn kho
            for item in noi_dung_phieu_cu:
                kq = misc.sql_one(f"SELECT * from ton_kho WHERE model = %s AND ma_kho = %s", (item[1], warehouse_name,))
                nhap = int(kq[6]) - int(item[2])
                ton = int(kq[4]) - int(item[2])

                misc.sql_commit(f"UPDATE ton_kho SET nhap = %s, ton = %s WHERE model = %s AND ma_kho = %s", (nhap, ton, item[1], warehouse_name,))

            sql_query = """
                        UPDATE nhap_kho SET ngay_thang = %s, noi_dung = %s, nguoi_lap = %s, ma_kho = %s, gio_nhap = %s,
                        WHERE id = %s
                        """
            params = (ngaythang, tb1, self.user, warehouse_name, current_time, sophieu)
            misc.sql_commit(sql_query, params)

            self.uic8.label_noti.setStyleSheet('color: blue')
            self.uic8.label_noti.setText('Đã sửa nội dung phiếu nhập kho')
            self.uic8.label_noti.repaint()

        else:
            sql_query = """
            INSERT INTO nhap_kho (id, ngay_thang, noi_dung, nguoi_lap, ma_kho, gio_nhap)
            VALUES (%s, %s, %s, %s, %s, %s)
            """
            params = (sophieu, ngaythang, tb1, self.user, warehouse_name, current_time,)
            misc.sql_commit(sql_query, params)

            self.uic8.label_noti.setStyleSheet('color: blue')
            self.uic8.label_noti.setText('Đã thêm mới phiếu nhập kho')
            self.uic8.label_noti.repaint()

        # Cộng số lượng nhập kho mới vào file ton_kho trên DB

        for item in tb:
            # Nếu chưa có model này trong kho này
            kq = misc.sql_one(f"SELECT * from ton_kho WHERE model = %s AND ma_kho = %s", (item[1], warehouse_name))
            if not kq:
                misc.sql_commit(
                    "INSERT INTO ton_kho (ten_san_pham, model, nhap, gia_nhap, ma_kho, ton) "
                    "VALUES (%s, %s, %s, %s, %s, %s)",
                    (item[0], item[1], item[2], item[3].replace(",", ""), warehouse_name, item[2])
                )

            else:
                sl_ton = int(kq[4])
                sl_da_nhap_kho = int(kq[6])
                gia_tb = int(kq[7])
                sl_nhap = int(item[2])
                gia_nhap_moi = int(item[3].replace(",", ""))
                gia_tb_moi = (sl_da_nhap_kho*gia_tb + sl_nhap*gia_nhap_moi)//(sl_da_nhap_kho + sl_nhap)

                misc.sql_commit(f"UPDATE ton_kho SET nhap = %s, ton = %s, gia_nhap = %s WHERE model = %s AND ma_kho = %s", (sl_nhap + sl_da_nhap_kho, sl_ton+sl_nhap, gia_tb_moi, item[1], warehouse_name))

    def luu_phieu_xuat(self):
        # try:
            warehouse_name = self.uic7.combo_kho.currentText()

            self.uic7.label_noti.clear()
            # Ghi nhận số liệu trên màn hình vào một biến gọi là hanghoa
            rows = self.uic7.tableWidget.rowCount()
            cols = self.uic7.tableWidget.columnCount()
            hanghoa = []
            for row in range(rows):
                if self.uic7.tableWidget.item(row, 2) is None or self.uic7.tableWidget.item(row, 3) is None:
                    pass
                else:
                    row_content = []
                    for col in range(cols):
                        item = self.uic7.tableWidget.item(row, col)
                        if item is not None:
                            if col == 3:
                                row_content.append(item.text().strip().replace(",", ""))
                            else:
                                row_content.append(item.text().strip())
                        else:
                            row_content.append('None')
                    # row_content.append(warehouse_name)
                    hanghoa.append(row_content)
            # Đã ghi xong hàng hóa
            hangthieu = ''
            list_xuat_kho = []

            for item in hanghoa:
                if item[1] in ['APP', 'NhanCong']:
                    item.append(warehouse_name)
                    item.append('OK')
                    list_xuat_kho.append(item)
                else:
                    done = False
                    sl_can = int(item[2])
                    # 'Kiểm tra số lượng trong kho đã chọn xem đủ hay không.  Nếu đủ thì xuất, ko thì tiếp tục kiểm tra các kho khác'
                    kq = misc.sql_one(f"SELECT * from ton_kho WHERE ma_kho = %s and model = %s", (warehouse_name, item[1],))
                    # Nếu trong kho có sẵn model này
                    if kq:
                        sl_co = kq[4]

                        if sl_co >= sl_can:
                            # Hàng này có sẵn trong kho - đủ số lượng
                            item1 = item[:]  # Create a copy
                            item1[4] = warehouse_name
                            item1.append('OK')
                            list_xuat_kho.append(item1)
                            sl_can = 0
                            done = True

                        elif 0 < sl_co < sl_can:
                            item1 = item[:]  # Create a copy
                            item1[2] = str(sl_co)
                            item1[4] = warehouse_name
                            item1.append('gom hàng')
                            list_xuat_kho.append(item1)
                            sl_can = sl_can - sl_co

                    if not done:  # Nếu không đủ thì đi gom hàng từ các kho khác
                        print('Kho ', warehouse_name, ' không có đủ ', item[1], '.')
                        allstock = ['Kho Hà Nội', 'Kho HCM']
                        allstock = [kho for kho in allstock if kho != warehouse_name]

                        for kho in allstock:

                            kq = misc.sql_one("SELECT * from ton_kho WHERE ma_kho = %s AND model = %s", (kho, item[1],))

                            if kq and done == False:
                                sl_co = kq[4]

                                if sl_co >= sl_can:
                                    item1 = item[:]  # Create a copy
                                    item1[2] = str(sl_can)
                                    try:
                                        item1[4] = kho
                                    except:
                                        item.append(kho)

                                    item1.append('OK')
                                    list_xuat_kho.append(item1)
                                    done = True
                                    sl_can = sl_can - sl_co

                                elif 0 < sl_co < sl_can:

                                    item1 = item[:]  # Create a copy
                                    item1[2] = str(sl_co)
                                    sl_can = sl_can - sl_co
                                    item1[4] = kho
                                    item1.append('OK')
                                    list_xuat_kho.append(item1)
                            else:
                                pass

                    if sl_can > 0:
                        hangthieu = hangthieu + ' ' + str(item[1])
                        item1 = item[:]  # Create a copy
                        item1[2] = str(sl_can)
                        try:
                            item1[4] = 'Không có hàng trong kho'
                        except:
                            item1.append('Không có hàng trong kho')
                        item1.append('Thiếu số này')
                        list_xuat_kho.append(item1)

            tien_hang = sum(int(ele[2])*int(ele[3]) for ele in hanghoa)
            self.uic7.text_gia_tri.setText("{:,}".format(tien_hang) + ' VNĐ')

            # Ghi nội dung biến tb - tức là nội dung phiếu xuất - lên màn hình
            self.uic7.tableWidget.clear()
            self.uic7.tableWidget.setColumnCount(5)
            self.uic7.tableWidget.setRowCount(len(list_xuat_kho))
            self.uic7.tableWidget.setHorizontalHeaderLabels(
                ['Tên sản phẩm', 'Model', 'Số lượng', 'Đơn giá', 'Kho hàng'])

            self.uic7.tableWidget.setColumnWidth(0, 380)
            self.uic7.tableWidget.setColumnWidth(1, 100)
            self.uic7.tableWidget.setColumnWidth(2, 80)
            self.uic7.tableWidget.setColumnWidth(3, 100)
            self.uic7.tableWidget.setColumnWidth(4, 100)

            for row in range(len(list_xuat_kho)):
                self.uic7.tableWidget.setItem(row, 0, QTableWidgetItem(str(list_xuat_kho[row][0])))
                self.uic7.tableWidget.setItem(row, 1, QTableWidgetItem(str(list_xuat_kho[row][1])))
                self.uic7.tableWidget.setItem(row, 2, QTableWidgetItem(str(list_xuat_kho[row][2])))
                self.uic7.tableWidget.setItem(row, 3, QTableWidgetItem("{:,}".format(int(list_xuat_kho[row][3]))))

                self.uic7.tableWidget.setItem(row, 4, QTableWidgetItem(str(list_xuat_kho[row][4])))
                if list_xuat_kho[row][4] == 'Không có hàng trong kho':  # Example condition to highlight rows
                    item = QTableWidgetItem(str(list_xuat_kho[row][4]))
                    item.setBackground(QColor(255, 0, 0))  # Set background color to red

            if hangthieu != '':
                self.uic7.label_noti.setStyleSheet('color: red')
                text = 'Model ' + hangthieu + ' không đủ hàng, kể cả khi gom các kho.'
                self.uic7.label_noti.setText(text)
                return
            else:
                nguoi_nhan = self.uic7.text_nguoi_nhan_hang.toPlainText().strip()
                sdt = self.uic7.text_sdt.toPlainText().strip()
                dia_chi = self.uic7.text_dia_chi.toPlainText().strip()
                try:
                    so_bg = int(self.uic7.text_so_bg.toPlainText().strip())
                except Exception as e:
                    print(e)
                    so_bg = 0
                try:
                    lead_id = int(self.uic7.text_lead.toPlainText().strip())
                except Exception as e:
                    lead_id = 0
                    print(e)

                if list_xuat_kho and nguoi_nhan and len(sdt) >= 10 and dia_chi:
                    nguoi_setup = self.uic7.combo_nguoi_thuc_hien.currentText()
                    if nguoi_setup == 'Người setup - giao nhận':
                        self.uic7.label_noti.setStyleSheet('color: red')
                        self.uic7.label_noti.setText('Chưa xác định người cài đặt thiết bị và giao hàng!')
                        self.uic7.label_noti.repaint()
                        return
                    else:
                        misc.sql_commit(f"UPDATE ds_don_hang SET nguoi_cai_dat = %s WHERE so_bg = %s", (nguoi_setup, so_bg,))

                        total_profit = 0

                        for item in list_xuat_kho:
                            profit = 0
                            # no = 0
                            if item != ['']:
                                gia_vao = [ele for ele in list(misc.sql_one("SELECT gia_dau_vao, gia_cap_1, gia_cap_2 FROM gia_tong_hop WHERE model = %s", (item[1],)))]

                                if item[1] == 'APP':
                                    gia_vao[0] = gia_vao[1] = gia_vao[2] = 360000
                                if item[1] == 'NhanCong':
                                    gia_vao[0] = gia_vao[1] = gia_vao[2] = round(int(item[3]) * 0.7)

                                profit = int(item[2]) * (int(item[3]) - int(gia_vao[0]))

                            # Tổng hợp
                            total_profit += profit

                        doanh_so = sum(int(item[3]) for item in list_xuat_kho)

                        total_profit = total_profit - doanh_so * 0.05

                        self.uic7.text_profit.setText("{:,}".format(round(total_profit)))

                        misc.sql_commit(f"UPDATE ds_don_hang SET profit = %s WHERE so_bg = %s", (total_profit, so_bg,))

                        StockHandle.xuat_kho(self, list_xuat_kho, nguoi_nhan, sdt, dia_chi, so_bg, lead_id)

                        self.uic7.label_noti.setStyleSheet('color: blue')
                        self.uic7.label_noti.setText('Đã gửi phiếu xuất kho này, đợi kế toán kho xuất hàng!')
                        self.uic7.label_noti.repaint()
                        self.uic7.tableWidget.repaint()
                        return
                else:
                    self.uic7.label_noti.setStyleSheet('color: red')
                    self.uic7.label_noti.setText('Vui lòng nhập đầy đủ thông tin của phiếu xuất kho - bao gồm cả địa chỉ nhận hàng.')
                    self.uic7.label_noti.repaint()
                    return

    def xuat_kho(self, list_xuat_kho, nguoi_nhan, sdt, dia_chi, so_bg, lead_id):
        # Lưu phiếu xuất kho vào DB
        sophieu = self.uic7.combo_sophieu.currentText()
        ngaythang = self.uic7.dateEdit.text()  # Ví dụ: '22/04/2025'
        dt_obj = datetime.strptime(ngaythang, "%d/%m/%Y")  # Nếu định dạng là dd/MM/yyyy
        ngaythang = dt_obj.strftime("%Y-%m-%d 00:00:00")

        makho = self.uic7.combo_kho.currentText()
        tieude = self.uic7.text_noi_dung_xuat.toPlainText().strip()
        if tieude == '': tieude = 'Báo giá thiết bị báo cháy FireSmart'

        doanh_so = sum(int(item[2])*int(item[3]) for item in list_xuat_kho)

        ds = []
        tb = []
        for item in list_xuat_kho:
            ds.append('|'.join(item))
            tb = "@@".join(ds)

        kq = misc.sql_one("SELECT * FROM xuat_kho WHERE id = %s", (sophieu,))
        # Nếu có số phiếu này rồi thì update nội dung phiếu
        if kq:
            noi_dung_phieu_cu = [ele.split('|') for ele in kq[2].split('@@')]

            for item in noi_dung_phieu_cu:
                if item[1] not in ['APP', 'NhanCong']:
                    kq = misc.sql_one("SELECT * from ton_kho WHERE model = %s AND ma_kho = %s", (item[1], item[4],))
                    xuat = int(kq[5]) - int(item[2])    # sửa tiếp ở đây
                    ton = int(kq[4]) + int(item[2])
                    misc.sql_commit(
                        f"UPDATE ton_kho SET xuat = %s, ton = %s WHERE model = %s AND ma_kho = %s", (xuat, ton, item[1], item[4],))

            sql_query = """
                                UPDATE xuat_kho SET ngay_thang = %s, noi_dung = %s, nguoi_lap = %s, ma_kho = %s, 
                                nguoi_nhan = %s, sdt = %s, address = %s, so_bg = %s, lead_id = %s 
                                WHERE id = %s
                                """
            params = (ngaythang, tb, self.user, makho, nguoi_nhan, sdt, dia_chi, so_bg, lead_id, sophieu)
            misc.sql_commit(sql_query, params)

        # Nếu chưa có số phiếu này thì tạo mới bản ghi và lưu nội dung
        else:
            sql_query = """
                    INSERT INTO xuat_kho (id, ngay_thang, noi_dung, nguoi_lap, ma_kho, nguoi_nhan, sdt, address, so_bg, lead_id, tieu_de, doanh_so)
                    VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
                    """
            params = (sophieu, ngaythang, tb, self.user, makho, nguoi_nhan, sdt, dia_chi, so_bg, lead_id, tieude, doanh_so)
            misc.sql_commit(sql_query, params)

        # Ghi sổ xuất nhập tồn
        for item in list_xuat_kho:
            if item[1] not in ['APP', 'NhanCong']:
                result = misc.sql_one("SELECT * from ton_kho WHERE model = %s AND ma_kho = %s", (item[1], item[4],))
                ton = int(result[4])
                xuat = int(result[5])
                ton_moi = ton - int(item[2])
                xuat_moi = xuat + int(item[2])

                sql_query = "UPDATE ton_kho SET ton = %s, xuat = %s WHERE model = %s AND ma_kho = %s "
                params = (ton_moi, xuat_moi, item[1], item[4])
                misc.sql_commit(sql_query, params)

        misc.sql_commit("UPDATE xuat_kho SET kt_duyet = 'F' WHERE so_bg = %s", (so_bg,))
        print('Đã chính thức xuất kho và ghi sổ')

    def khoi_tao(self, uic):
        table = uic.tableWidget
        stock_ui_utils.setup_table_xuat(table, 28)

        # === LOAD SẢN PHẨM TỪ DB ===
        product_rows = misc.sql_all(
            "SELECT model, ten_san_pham FROM gia_tong_hop",
            None
        )

        # === CACHE ĐÃ CHUẨN HOÁ ===
        model_to_name = {}
        name_to_model = {}

        for row in product_rows:
            raw_model = row[0]
            raw_name = row[1]

            model_key = _normalize_key(raw_model)
            name_key = _normalize_key(raw_name)

            if model_key:
                model_to_name[model_key] = raw_name
            if name_key:
                name_to_model[name_key] = raw_model

        # === GẮN AUTOCOMPLETE CHO TỪNG DÒNG ===
        for row_idx in range(table.rowCount()):
            setup_autocomplete_for_table_row(
                table,
                row_idx,
                model_to_name,
                name_to_model
            )

    def xem_ton_kho(self, tenkho):
        # StockHandle.tong_ket_kho(self)
        self.uic7.but_xuat_kho.setHidden(True)
        self.uic7.but_themdong.setDisabled(True)
        self.uic7.but_xoadong.setDisabled(True)
        self.uic7.but_huy_phieu_xuat.setHidden(True)
        self.uic7.group_xuatkho.hide()

        # self.uic7.combo_sophieu.clear()
        self.uic7.label_noti.clear()
        self.uic7.label_noti.repaint()
        self.uic7.label.setText('Số lượng hàng tồn kho')
        self.uic7.label.repaint()

        self.uic7.tableWidget.clearContents()

        result = misc.sql_all("SELECT * FROM ton_kho WHERE ma_kho = %s", (tenkho,))

        if result:
            stock_ui_utils.setup_table_xuat(self.uic7.tableWidget, len(result))
            self.uic7.tableWidget.setHorizontalHeaderLabels(['Tên sản phẩm', 'Model', 'tồn kho', 'Giá nhập kho', 'Mã kho'])


            for row in range(len(result)):
                self.uic7.tableWidget.setItem(row, 0, QTableWidgetItem(result[row][1]))
                self.uic7.tableWidget.setItem(row, 1, QTableWidgetItem(result[row][2]))
                self.uic7.tableWidget.setItem(row, 2, QTableWidgetItem(str(result[row][4])))
                self.uic7.tableWidget.setItem(row, 3, QTableWidgetItem(str(result[row][7])))
                self.uic7.tableWidget.setItem(row, 4, QTableWidgetItem(str(result[row][8])))
        self.uic7.tableWidget.repaint()

    def xoa_dong(self, uic):
        table = uic.tableWidget
        current_row = table.currentRow()

        if current_row >= 0:
            table.removeRow(current_row)
            uic.label_noti.setStyleSheet('color: red')
            uic.label_noti.setText(f'⚠️ Đã xóa dòng thứ {current_row + 1}.')
        else:
            uic.label_noti.setStyleSheet('color: red')
            uic.label_noti.setText('⚠️ Click vào dòng cần xóa trước đã.')


    def them_dong(self, uic):
        table = uic.tableWidget
        current_row_count = table.rowCount()
        table.insertRow(current_row_count)

        # Lấy dữ liệu ánh xạ từ CSDL
        product_rows = misc.sql_all("SELECT model, ten_san_pham FROM gia_tong_hop", None)
        model_to_name = {row[0]: row[1] for row in product_rows}
        name_to_model = {row[1]: row[0] for row in product_rows}

        # Gọi hàm gắn auto-complete cho dòng mới
        setup_autocomplete_for_table_row(table, current_row_count, model_to_name, name_to_model)

    def xuat_file(self, sophieu):
        self.uic7.but_huy_phieu_xuat.setHidden(True)
        # Connect to SQL Server database

        kq = misc.sql_one("SELECT * FROM xuat_kho WHERE id = %s", (sophieu,))
        noidungxuat = kq[2] # sẽ ghi vào file ds_don_hang ở cuối def này

        hanghoa = [ele.split('|') for ele in kq[2].split("@@")]

        workbook = openpyxl.load_workbook('mau_phieu_xuat_kho.xlsx')

        # Access the specified worksheet
        worksheet = workbook['Sheet1']
        sheet = worksheet
        cell = worksheet.cell(row=13, column=1)
        cell.value = 'Người lập phiếu: ' + kq[3]

        cell = worksheet.cell(row=8, column=3)
        cell.value = kq[9]  # Tên khách hàng

        cell = worksheet.cell(row=9, column=3)
        cell.value = kq[10]

        cell = worksheet.cell(row=10, column=3)
        cell.value = kq[11]

        cell = worksheet.cell(row=5, column=1)
        date_str = str(datetime.now().date())
        date_obj = datetime.strptime(date_str, '%Y-%m-%d')

        # Chuyển đổi đối tượng datetime thành chuỗi mới trong định dạng "DD-MM-YYYY"
        new_date_str = date_obj.strftime('%d-%m-%Y')
        cell.value = 'Ngày: ' + str(new_date_str)

        cell = worksheet.cell(row=6, column=1)
        cell.value = 'Số: PXK-' + str(sophieu)

        cell = worksheet.cell(row=12, column=3)
        cell.value = kq[8]
        #
        for i in range(len(hanghoa)+1):
            try:
                # des[i].append(str(int(des[i][5]) * int(des[i][6].replace(",", ""))))
                cell = worksheet.cell(row=16 + i, column=1)
                cell.value = i+1

                cell = worksheet.cell(row=16 + i, column=2)
                cell.value = hanghoa[i][1]

                cell = worksheet.cell(row=16 + i, column=3)
                cell.value = hanghoa[i][0]

                cell = worksheet.cell(row=16 + i, column=4)
                cell.value = 'cái'

                cell = worksheet.cell(row=16 + i, column=5)
                cell.value = hanghoa[i][2]

                cell = worksheet.cell(row=16 + i, column=6)
                cell.number_format = "#,##0"
                cell.value = int(hanghoa[i][3])

                cell = worksheet.cell(row=16 + i, column=7)
                cell.number_format = "#,##0"
                cell.value = f'=E{16+i}*F{16+i}'
            except:
                pass

        worksheet.delete_rows(16 + len(hanghoa), 70 - len(hanghoa))
        # Ghi tổng trước thuế:
        cell = worksheet.cell(row=18 + len(hanghoa), column=7)
        cell.number_format = "#,##0"
        cell.value = f'=sum(G16:G{16+len(hanghoa)})'

        try:
            mst = misc.sql_one("SELECT mst FROM sale_lead WHERE lead_id = %s", (kq[7],))[0]
            if not mst:
                mst = '000'
        except:
            mst = '000'

        if len(mst) > 9:
            # Ghi thuế 8
            cell = worksheet.cell(row=19 + len(hanghoa), column=7)
            cell.number_format = "#,##0"
            cell.value = f'=G{18+len(hanghoa)}*0.08'
        else:
            # Ghi thuế 8
            cell = worksheet.cell(row=19 + len(hanghoa), column=7)
            cell.number_format = "#,##0"
            cell.value = 0

        # Tổng sau thuế
        cell = worksheet.cell(row=21 + len(hanghoa), column=7)
        cell.number_format = "#,##0"
        cell.value = f'=G{18+len(hanghoa)}+G{19+len(hanghoa)}+G{20+len(hanghoa)}'

        # Số tiền bằng chữ

        tong_truoc_thue = sum(int(ele[2])*int(ele[3]) for ele in hanghoa)
        if len(mst) > 9:
            vat = tong_truoc_thue*0.08
        else:
            vat = 0
        tong = round(tong_truoc_thue + vat)

        cell = worksheet.cell(row=22 + len(hanghoa), column=3)
        cell.number_format = "#,##0"
        cell.value = StockHandle.num_to_vietnamese(self, tong)

        # Ghi ngày tháng năm
        cell = worksheet.cell(row=24 + len(hanghoa), column=5)
        cell.number_format = "#,##0"
        cell.value = 'Ngày: ' + str(new_date_str)

        # Create a tkinter root window (it won't be displayed)
        root = tk.Tk()
        root.withdraw()

        file_save_path = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel Files", "*.xlsx")],
            initialfile=f"PXK-{sophieu}----.xlsx"  # <- Set your default file name here
        )

        # Check if the user canceled the save dialog
        if not file_save_path:
            tex = 'Chưa SAVE file excel.'
            self.uic7.label_noti.setStyleSheet('color: red')
        else:
            try:
                # Save the modified workbook to the selected location
                workbook.save(file_save_path)

                # Find the last occurrence of '/'
                last_slash = file_save_path.rfind('/')

                # Extract substring from last slash to the end
                if last_slash != -1:  # Check if '/' is found
                    tenfile = file_save_path[last_slash + 1:]  # +1 to exclude the slash itself
                else:
                    tenfile = file_save_path  # If '/' is not found, return the original string

                tex = "Đã SAVE file:   " + tenfile
                self.uic7.label_noti.setStyleSheet('color: blue')

                text = 'Kế toán đã xuất kho theo phiếu xuất số: ' + str(sophieu) + ' của ' + str(kq[3])
                misc.send_to_telegram(text)

                so_bg = kq[6]
                # Đánh dấu là kế toán đã xuất kho phiếu xuất sophieu này
                current_time = datetime.now().strftime("%H:%M")
                misc.sql_commit("UPDATE xuat_kho SET kt_duyet = 'T', nguoi_duyet = %s, gio_duyet = %s WHERE id = %s", (self.user, current_time, sophieu,))

                ngaythang = datetime.now().strftime('%Y-%m-%d')
                misc.sql_commit("UPDATE ds_don_hang SET ngay_hen_giao_hang = %s WHERE so_bg = %s", (ngaythang, so_bg,))
                misc.sql_commit("UPDATE ds_don_thue SET ngay_ban_giao = %s WHERE so_bg = %s", (ngaythang, so_bg,))

                # Đánh dấu là đơn hàng bán đã hoàn thành
                misc.sql_commit("UPDATE ds_don_hang SET da_hoan_thanh = 'T', ghi_chu = %s WHERE so_bg = %s", (noidungxuat, so_bg,))
                # Đánh dấu là đơn hàng thuê đã hoàn thành (NẾU LÀ THUÊ)
                misc.sql_commit("UPDATE ds_don_thue SET hoan_thanh = 'T' WHERE so_bg = %s", (so_bg,))
                # Đánh dấu table ds_bao_gia là báo giá đã thành công, ko sửa được nữa
                misc.sql_commit("UPDATE ds_bao_gia SET thanh_cong = 'T' WHERE so_bg = %s", (so_bg,))

                lead_row = misc.sql_one("SELECT lead_id, status FROM sale_lead WHERE dat_hang='T' AND lead_id IN (SELECT lead_id FROM ds_don_hang WHERE so_bg=%s)", (so_bg,))
                if lead_row:
                    lid, old_status = lead_row[0], lead_row[1]
                    if str(old_status or '') != 'Đã giao hàng':
                        misc.sql_commit("UPDATE sale_lead SET status='Đã giao hàng' WHERE lead_id=%s", (lid,))
                        misc.refresh_busy_for_lead(lid)
                        misc.audit_log(self.user, 'EXPORT_STOCK', 'status', old_status, 'Đã giao hàng', lid)
                misc.audit_log(self.user, 'EXPORT_STOCK', 'so_bg', '-', so_bg, lead_row[0] if lead_row else None)

            except Exception as e:
                print(e)
                tex = 'Có lỗi khi save file.'

        self.uic7.label_noti.setText(tex)
        self.uic7.label_noti.repaint()
        # Close the tkinter root window
        root.destroy()

    def num_to_vietnamese(self, n):
        try:
            units = ["không", "một", "hai", "ba", "bốn", "năm", "sáu", "bảy", "tám", "chín"]
            tens = ["", "mười", "hai mươi", "ba mươi", "bốn mươi", "năm mươi", "sáu mươi", "bảy mươi", "tám mươi",
                    "chín mươi"]
            scales = ["", "nghìn", "triệu", "tỷ"]

            if n == 0:
                return "Không"

            def three_digit_to_words(n):
                hundred = n // 100
                ten = (n // 10) % 10
                unit = n % 10
                words = []

                if hundred:
                    words.append(units[hundred] + " trăm")
                    if ten == 0 and unit:
                        words.append("lẻ")

                if ten:
                    words.append(tens[ten])
                elif hundred and unit:
                    words.append("")

                if unit:
                    if ten == 1 and unit == 5:
                        words.append("lăm")
                    else:
                        words.append(units[unit])
                return " ".join(words).strip()

            words = []
            scale_idx = 0

            while n > 0:
                chunk = n % 1000
                if chunk:
                    words.append(scales[scale_idx])
                    words.append(three_digit_to_words(chunk))
                n //= 1000
                scale_idx += 1

            words = list(filter(None, words))
            words.reverse()
            result = " ".join(words).strip()
            return result[0].upper() + result[1:] + ' VNĐ.'
        except Exception as e:
            print(e)
            return ''

    def xuat_kho_thue(self, so_bg):

        try:
            self.win_cho_thue.close()
        except:
            pass

        self.win_xuat_thue = QMainWindow()
        self.uic14 = Ui_xuat_kho_thue()
        self.uic14.setupUi(self.win_xuat_thue)
        apply_ui_v2(self.win_xuat_thue)
        self.win_xuat_thue.show()

        self.uic14.combo_kho.addItems(["Kho Hà Nội", "Kho HCM"])
        self.uic14.text_so_bg.setText(str(so_bg))

        kq = misc.sql_one("SELECT * FROM ds_don_thue WHERE so_bg = %s", (so_bg,))

        gia_tri = int(kq[2]) + int(kq[3]) + int(kq[5])
        self.uic14.text_tong_cong.setText("{:,}".format(gia_tri))
        self.uic14.text_da_thanhtoan.setText("{:,}".format(kq[6]))
        self.uic14.text_phai_thu.setText("{:,}".format(gia_tri - int(kq[6])))
        self.uic14.text_thu_theo_ky.setText("{:,}".format(kq[2]))
        self.uic14.text_dat_coc.setText("{:,}".format(kq[5]))

        kq = misc.sql_one("SELECT * FROM sale_lead WHERE lead_id = %s", (kq[1],))
        self.uic14.text_nguoi_nhan_hang.setText(kq[1])
        self.uic14.text_sdt.setText(kq[2])
        self.uic14.text_dia_chi.setText(kq[12])

        kq = misc.sql_one('SELECT * from xuat_kho WHERE so_bg = %s', (so_bg,))

        if kq:
            # Lấy số phiếu xuất
            so_phieu_xuat = kq[0]
            self.uic14.combo_sophieu.clear()
            self.uic14.combo_sophieu.addItems([so_phieu_xuat])

        else:
            # Lấy số phiếu xuất kho mới
            kq = misc.sql_one("SELECT * FROM xuat_kho WHERE id = (SELECT MAX(id) FROM xuat_kho)")
            try:
                self.uic14.combo_sophieu.clear()
                self.uic14.combo_sophieu.addItems([str(kq[0] + 1)])
            except Exception as e:
                print(e)

        # Ghi nội dung biến hanghoa - tức là nội dung phiếu xuất - lên màn hình
        kq = misc.sql_one("SELECT noi_dung FROM ds_bao_gia WHERE so_bg = %s", (so_bg,))[0].split('@')
        hanghoa = []
        for item in kq:
            hanghoa.append(item.split('|'))
        self.uic14.tableWidget.clear()
        self.uic14.tableWidget.setColumnCount(5)
        self.uic14.tableWidget.setRowCount(len(hanghoa))
        self.uic14.tableWidget.setHorizontalHeaderLabels(
            ['Tên sản phẩm', 'Model', 'Số lượng', 'Đơn giá', 'Kho hàng'])

        self.uic14.tableWidget.setColumnWidth(0, 380)
        self.uic14.tableWidget.setColumnWidth(1, 100)
        self.uic14.tableWidget.setColumnWidth(2, 80)
        self.uic14.tableWidget.setColumnWidth(3, 100)
        self.uic14.tableWidget.setColumnWidth(4, 100)

        for row in range(len(hanghoa)):
            if hanghoa[row] == ['']:
                pass
            # elif str(hanghoa[row][1]) != 'NhanCong':
            else:
                self.uic14.tableWidget.setItem(row, 0, QTableWidgetItem(str(hanghoa[row][0])))
                self.uic14.tableWidget.setItem(row, 1, QTableWidgetItem(str(hanghoa[row][1])))
                self.uic14.tableWidget.setItem(row, 2, QTableWidgetItem(str(hanghoa[row][4])))
                self.uic14.tableWidget.setItem(row, 3,
                                               QTableWidgetItem("{:,}".format(int(hanghoa[row][5].replace(",", "")))))

        self.uic14.tableWidget.repaint()
        self.uic14.tableWidget.setEditTriggers(QAbstractItemView.EditTrigger.NoEditTriggers)

        kq = misc.sql_all("SELECT full_name FROM user WHERE power > 0 and power < 50", None)
        # Lấy danh sách user điền vào ô người giao nhận
        kq1 = list(kq)

        self.uic14.combo_nguoi_thuc_hien.addItem('Người setup - giao nhận')
        self.uic14.combo_nguoi_thuc_hien.addItems(item[0] for item in kq1)

        self.uic14.but_save_phieu_xuat.clicked.connect(lambda: StockHandle.save_phieu_xuat_thue(self))

    def save_phieu_xuat_thue(self):
        ncd = self.uic14.combo_nguoi_thuc_hien.currentText()
        if ncd == 'Người setup - giao nhận':
            self.uic14.label_noti.setStyleSheet('color: red')
            self.uic14.label_noti.setText('Cần phải chỉ định người cài đặt và giao nhận!')
            return
        else:
            self.uic14.label_noti.clear()

        warehouse_name = self.uic14.combo_kho.currentText()

        # Ghi nhận số liệu trên màn hình vào một biến gọi là hanghoa
        rows = self.uic14.tableWidget.rowCount()
        cols = self.uic14.tableWidget.columnCount()
        hanghoa = []
        for row in range(rows):
            if self.uic14.tableWidget.item(row, 2) is None or self.uic14.tableWidget.item(row, 3) is None:
                pass
            else:
                row_content = []
                for col in range(cols):
                    item = self.uic14.tableWidget.item(row, col)
                    if item is not None:
                        if col == 3:
                            row_content.append(item.text().strip().replace(",", ""))
                        else:
                            row_content.append(item.text().strip())
                    else:
                        row_content.append('None')
                # row_content.append(warehouse_name)
                hanghoa.append(row_content)
        # Đã ghi xong hàng hóa
        hangthieu = ''
        list_xuat_kho = []

        for item in hanghoa:
            if item[1] in ['APP', 'NhanCong']:
                item.append(warehouse_name)
                item.append('OK')
                list_xuat_kho.append(item)
            else:
                done = False
                sl_can = int(item[2])
                # 'Kiểm tra số lượng trong kho đã chọn xem đủ hay không.  Nếu đủ thì xuất, ko thì tiếp tục kiểm tra các kho khác'
                kq = misc.sql_one(f"SELECT * from ton_kho WHERE ma_kho = %s and model = %s", (warehouse_name, item[1],))
                # Nếu trong kho có sẵn model này
                if kq:
                    sl_co = kq[4]

                    if sl_co >= sl_can:
                        # Hàng này có sẵn trong kho - đủ số lượng
                        item1 = item[:]  # Create a copy
                        item1[4] = warehouse_name
                        item1.append('OK')
                        list_xuat_kho.append(item1)
                        sl_can = 0
                        done = True

                    elif 0 < sl_co < sl_can:
                        item1 = item[:]  # Create a copy
                        item1[2] = str(sl_co)
                        item1[4] = warehouse_name
                        item1.append('gom hàng')
                        list_xuat_kho.append(item1)
                        sl_can = sl_can - sl_co

                if not done:  # Nếu không đủ thì đi gom hàng từ các kho khác
                    print('Kho ', warehouse_name, ' không có đủ ', item[1], '.')
                    allstock = ['Kho Hà Nội', 'Kho HCM']
                    allstock = [kho for kho in allstock if kho != warehouse_name]

                    for kho in allstock:

                        kq = misc.sql_one(f"SELECT * from ton_kho WHERE ma_kho = %s and model = %s", (kho, item[1],))

                        if kq and done == False:
                            sl_co = kq[4]

                            if sl_co >= sl_can:
                                item1 = item[:]  # Create a copy
                                item1[2] = str(sl_can)
                                try:
                                    item1[4] = kho
                                except:
                                    item.append(kho)

                                item1.append('OK')
                                list_xuat_kho.append(item1)
                                done = True
                                sl_can = sl_can - sl_co

                            elif 0 < sl_co < sl_can:

                                item1 = item[:]  # Create a copy
                                item1[2] = str(sl_co)
                                sl_can = sl_can - sl_co
                                item1[4] = kho
                                item1.append('OK')
                                list_xuat_kho.append(item1)
                        else:
                            pass

                if sl_can > 0:
                    hangthieu = hangthieu + ' ' + str(item[1])
                    item1 = item[:]  # Create a copy
                    item1[2] = str(sl_can)
                    try:
                        item1[4] = 'Không có hàng trong kho'
                    except:
                        item1.append('Không có hàng trong kho')
                    item1.append('Thiếu số này')
                    list_xuat_kho.append(item1)

        # Ghi nội dung biến tb - tức là nội dung phiếu xuất - lên màn hình

        for row in range(len(list_xuat_kho)):
            self.uic14.tableWidget.setItem(row, 4, QTableWidgetItem(str(list_xuat_kho[row][4])))
            if list_xuat_kho[row][4] == 'Không có hàng trong kho':  # Example condition to highlight rows
                item = QTableWidgetItem(str(list_xuat_kho[row][4]))
                item.setBackground(QColor(255, 0, 0))  # Set background color to red

        if hangthieu != '':
            self.uic14.label_noti.setStyleSheet('color: red')
            text = 'Model ' + hangthieu + ' không đủ hàng, kể cả khi gom các kho.'
            self.uic14.label_noti.setText(text)
            return
        else:
            nguoi_nhan = self.uic14.text_nguoi_nhan_hang.toPlainText().strip()
            sdt = self.uic14.text_sdt.toPlainText().strip()
            dia_chi = self.uic14.text_dia_chi.toPlainText().strip()
            so_bg = int(self.uic14.text_so_bg.toPlainText().strip())

            if list_xuat_kho and nguoi_nhan and len(sdt) >= 10 and dia_chi:
                nguoi_setup = self.uic14.combo_nguoi_thuc_hien.currentText()

                misc.sql_commit(f"UPDATE ds_don_thue SET nguoi_cai_dat = %s WHERE so_bg = %s", (nguoi_setup, so_bg,))

                total_profit = 0

                for item in list_xuat_kho:
                    profit = 0
                    # no = 0
                    if item != ['']:
                        gia_vao = [ele for ele in list(misc.sql_one("SELECT gia_dau_vao, gia_thue FROM gia_tong_hop WHERE model = %s", (item[1],)))]

                        if item[1] == 'APP':
                            gia_vao[0] = gia_vao[1] = gia_vao[2] = 360000
                        if item[1] == 'NhanCong':
                            gia_vao[0] = gia_vao[1] = gia_vao[2] = round(int(item[3]) * 0.7)

                        profit = int(item[2]) * (int(item[3]) - int(gia_vao[0]))

                    # Tổng hợp
                    total_profit += profit

                misc.sql_commit("UPDATE ds_don_thue SET profit = %s WHERE so_bg = %s", (total_profit, so_bg,))
                lead_id = misc.sql_one("SELECT lead_id FROM ds_bao_gia WHERE so_bg = %s", (so_bg,))[0]

                StockHandle.save_phieu_xuat_thue_part_2(self, list_xuat_kho, nguoi_nhan, sdt, dia_chi, so_bg, lead_id)

                self.uic14.label_noti.setStyleSheet('color: blue')
                self.uic14.label_noti.setText('Đã gửi phiếu xuất kho này, đợi kế toán kho xuất hàng!')
                self.uic14.label_noti.repaint()
                self.uic14.tableWidget.repaint()

                return
            else:
                self.uic7.label_noti.setStyleSheet('color: red')
                self.uic7.label_noti.setText('Vui lòng nhập đầy đủ thông tin của phiếu xuất kho - bao gồm cả địa chỉ nhận hàng.')
                self.uic7.label_noti.repaint()
                return

    def save_phieu_xuat_thue_part_2(self, list_xuat_kho, nguoi_nhan, sdt, dia_chi, so_bg, lead_id):
        # Lưu phiếu xuất kho vào DB
        sophieu = self.uic14.combo_sophieu.currentText()
        ngaythang = self.uic14.dateEdit.text()
        makho = self.uic14.combo_kho.currentText()
        tieude = self.uic14.text_noi_dung_xuat.toPlainText().strip()
        if tieude == '': tieude = 'Báo giá thiết bị báo cháy FireSmart'

        doanh_so = sum(int(item[2])*int(item[3]) for item in list_xuat_kho)

        ds = []
        tb = []
        for item in list_xuat_kho:
            ds.append('|'.join(item))
            tb = "@@".join(ds)

        kq = misc.sql_one("SELECT * FROM xuat_kho WHERE id = %s", (sophieu,))

        # Nếu có số phiếu này rồi thì update nội dung phiếu
        if kq:
            noi_dung_phieu_cu = [ele.split('|') for ele in kq[2].split('@@')]

            for item in noi_dung_phieu_cu:
                if item[1] not in ['APP', 'NhanCong']:
                    kq = misc.sql_one("SELECT * from ton_kho WHERE model = %s AND ma_kho = %s", (item[1], item[4],))
                    xuat = int(kq[5]) - int(item[2])    # sửa tiếp ở đây
                    ton = int(kq[4]) + int(item[2])
                    misc.sql_commit(
                        f"UPDATE ton_kho SET xuat = %s, ton = %s WHERE model = %s AND ma_kho = %s", (xuat, ton, item[1], item[4],))

            sql_query = """
                                UPDATE xuat_kho SET ngay_thang = %s, noi_dung = %s, nguoi_lap = %s, ma_kho = %s, 
                                nguoi_nhan = %s, sdt = %s, address = %s, so_bg = %s, lead_id = %s 
                                WHERE id = %s
                                """
            params = (ngaythang, tb, self.user, makho, nguoi_nhan, sdt, dia_chi, so_bg, lead_id, sophieu)
            misc.sql_commit(sql_query, params)

        # Nếu chưa có số phiếu này thì tạo mới bản ghi và lưu nội dung
        else:
            sql_query = """
                    INSERT INTO xuat_kho (id, ngay_thang, noi_dung, nguoi_lap, ma_kho, nguoi_nhan, sdt, address, so_bg, lead_id, tieu_de, doanh_so)
                    VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
                    """
            params = (sophieu, ngaythang, tb, self.user, makho, nguoi_nhan, sdt, dia_chi, so_bg, lead_id, tieude, doanh_so)
            misc.sql_commit(sql_query, params)

        # Ghi sổ xuất nhập tồn
        for item in list_xuat_kho:
            if item[1] not in ['APP', 'NhanCong']:
                result = misc.sql_one("SELECT * from ton_kho WHERE model = %s AND ma_kho = %s", (item[1], item[4],))
                ton = int(result[4])
                xuat = int(result[5])
                ton_moi = ton - int(item[2])
                xuat_moi = xuat + int(item[2])

                print('Tồn: ', ton)
                print('Xuất: ', xuat)
                print('Tồn mới:', ton_moi)
                print('Xuất mơi: ', xuat_moi)

                sql_query = "UPDATE ton_kho SET ton = %s, xuat = %s WHERE model = %s AND ma_kho = %s "
                params = (ton_moi, xuat_moi, item[1], item[4])
                misc.sql_commit(sql_query, params)

        misc.sql_commit("UPDATE xuat_kho SET kt_duyet = 'F' WHERE so_bg = %s", (so_bg,))
        print('Đã chính thức xuất kho và ghi sổ')
