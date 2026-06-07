"""
tu_van_pccc_excel.py - Xuat bao gia PCCC theo template bao_gia_mau.xlsx.

Tach rieng de de test. Khong phu thuoc Qt, chi dung openpyxl.
Khac ham misc.save_excel() o cho: nhan truc tiep danh sach bao gia
+ thong tin khach hang (KHONG doc tu DB ds_bao_gia), vi module Tu van PCCC
khong luu bao gia vao DB.
"""

import os
import re
import unicodedata
import datetime
import openpyxl

from misc import get_resource_path


def _clean_filename(name: str) -> str:
    name = unicodedata.normalize("NFKD", name)
    name = "".join(c for c in name if c.isprintable())
    name = re.sub(r"\s+", " ", name)
    name = re.sub(r'[\\/*?:"<>|]', "", name)
    return name.strip()[:120]


def xuat_bao_gia_pccc(
    bg_rows: list,
    ttkh: dict,
    nguoi_lap: str,
    sdt_nguoi_lap: str,
    file_out: str,
    template: str = "bao_gia_mau.xlsx",
) -> str:
    """
    Tao file Excel bao gia PCCC tu template cong ty.

    Args:
        bg_rows: list cac dong bao gia, moi item la dict:
            {ten, model, hieu, dv, sl, gia, vat}
        ttkh: dict {ten, dia_chi, sdt, vv}
        nguoi_lap: ten nguoi lap bao gia
        sdt_nguoi_lap: SDT nguoi lap
        file_out: duong dan file output
        template: ten file template (mac dinh bao_gia_mau.xlsx)

    Returns: duong dan file da luu.
    """
    template_path = get_resource_path(template)
    if not os.path.exists(template_path):
        raise FileNotFoundError(
            "Khong tim thay file mau bao gia: " + template_path)

    wb = openpyxl.load_workbook(template_path)
    sheet = wb["Quotation"]

    now_str = datetime.datetime.now().strftime("%d-%m-%Y")
    so_bg = "PCCC-" + datetime.datetime.now().strftime("%y%m%d%H%M%S")

    # --- Thong tin khach hang ---
    sheet.cell(row=5, column=2).value = "Kinh gui: " + str(ttkh.get("ten", ""))
    sheet.cell(row=6, column=2).value = "Anh/Chi: " + str(ttkh.get("ten", ""))
    sheet.cell(row=7, column=2).value = "Dia chi: " + str(ttkh.get("dia_chi", ""))
    sheet.cell(row=8, column=2).value = "Dien thoai: " + str(ttkh.get("sdt", ""))
    sheet.cell(row=5, column=9).value = now_str
    sheet.cell(row=6, column=9).value = so_bg
    vv = ttkh.get("vv") or "Cung cap lap dat he thong PCCC"
    sheet.cell(row=9, column=2).value = "V/v: " + str(vv)

    # --- Du lieu hang hoa: bat dau tu row 12 ---
    # Theo dung pattern misc.save_excel(): label tong/VAT o cot 3 do TEMPLATE
    # cung cap san o rows 93-96. Sau delete_rows(13+N, 80-N) chung dich len
    # rows 13+N .. 16+N. Minh chi ghi gia tri vao cot 9 (I).
    sum8 = 0
    sum10 = 0
    for n, r in enumerate(bg_rows):
        row = 12 + n
        stt = str(n + 1)
        vat = int(r.get("vat", 8) or 8)
        prefix = {8: ".", 10: ".."}.get(vat, "")

        sl = float(r.get("sl") or 0)
        gia = float(r.get("gia") or 0)
        tt = sl * gia
        if vat == 10:
            sum10 += tt
        else:
            sum8 += tt

        sheet.cell(row=row, column=2).value = stt + prefix
        sheet.cell(row=row, column=3).value = r.get("ten", "")
        sheet.cell(row=row, column=4).value = r.get("model", "") or ""
        sheet.cell(row=row, column=5).value = r.get("hieu", "") or ""
        sheet.cell(row=row, column=6).value = r.get("dv", "") or ""
        sheet.cell(row=row, column=7).value = sl
        cell_gia = sheet.cell(row=row, column=8)
        cell_gia.value = gia
        cell_gia.number_format = "#,##0"
        # I (cot 9) da co san cong thuc =G*H trong template

    n_rows = len(bg_rows)
    last_data_row = 11 + n_rows

    # --- Xoa cac dong formula thua ---
    if n_rows < 80:
        sheet.delete_rows(13 + n_rows, 80 - n_rows)

    # --- Sau delete_rows: rows 13+N..16+N chua label tong/VAT tu template ---
    row_cong = 13 + n_rows
    row_vat8 = 14 + n_rows
    row_vat10 = 15 + n_rows
    row_tong = 16 + n_rows

    sheet.cell(row=row_cong, column=9).value = (
        "=SUM(I12:I" + str(last_data_row) + ")")
    sheet.cell(row=row_vat8, column=9).value = round(sum8 * 0.08)
    sheet.cell(row=row_vat10, column=9).value = round(sum10 * 0.1)
    sheet.cell(row=row_tong, column=9).value = (
        "=I" + str(row_cong) + "+I" + str(row_vat8) + "+I" + str(row_vat10))

    for rr in (row_cong, row_vat8, row_vat10, row_tong):
        sheet.cell(row=rr, column=9).number_format = "#,##0"

    # --- Nguoi bao gia (cot H, rows 22+N va 23+N giong save_excel) ---
    sheet.cell(row=22 + n_rows, column=8).value = nguoi_lap
    sheet.cell(row=23 + n_rows, column=8).value = sdt_nguoi_lap

    # --- Luu ---
    if not file_out.lower().endswith(".xlsx"):
        file_out = file_out + ".xlsx"
    wb.save(file_out)
    return file_out


def goi_y_ten_file(ten_kh: str, vv: str) -> str:
    """Sinh ten file mac dinh cho dialog Save As."""
    kh = ten_kh or "KH"
    v = vv or "He thong PCCC"
    return _clean_filename("BG PCCC - " + kh + " - " + v) + ".xlsx"
