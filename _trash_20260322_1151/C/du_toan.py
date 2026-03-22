import sys
from PyQt6.QtWidgets import (
    QApplication, QMainWindow, QFileDialog, QMessageBox, QTableWidgetItem, QComboBox, QWidget, QHBoxLayout,
    QTableWidget, QPushButton
)
from PyQt6.QtCore import Qt
from PyQt6.QtWidgets import QAbstractItemView

from openpyxl import load_workbook
from difflib import SequenceMatcher

import misc
from UI.win_du_toan import Ui_DuToan


class MainWindow(QMainWindow):
    def __init__(self):
        super().__init__()
        self.uic = Ui_DuToan()
        self.uic.setupUi(self)
        self.setWindowTitle(QApplication.translate("MainWindow", "Fsale v2.1.1"))

        # Map tab indices to table names
        self.table_widget_map = {
            0: self.uic.tableWidget,
            1: self.uic.tableWidget_2,
            2: self.uic.tableWidget_3,
            3: self.uic.tableWidget_4,
            4: self.uic.tableWidget_5,
            5: self.uic.tableWidget_6,
        }

        # Enable drag and drop for all table widgets
        for table_widget in self.table_widget_map.values():
            self.enable_drag_and_drop(table_widget)

        # Connect the buttons to their methods
        self.uic.but_load_file.clicked.connect(self.open_excel_file)
        self.uic.but_generate.clicked.connect(self.merge_columns_with_same_combobox_text)

        self.uic.but_xoa_cot.clicked.connect(self.delete_current_column)
        self.uic.but_xoa_dong.clicked.connect(self.delete_current_row)
        self.uic.but_them_dong.clicked.connect(self.add_new_row)
        self.uic.but_them_cot.clicked.connect(self.add_new_column)

    def enable_drag_and_drop(self, table_widget):
        """Enable drag-and-drop for columns in the given QTableWidget."""
        header = table_widget.horizontalHeader()
        header.setSectionsMovable(True)  # Allow columns to be reordered
        header.setDragEnabled(True)  # Enable drag functionality
        header.setDragDropMode(QAbstractItemView.DragDropMode.InternalMove)  # Enable internal movement
        header.setDefaultAlignment(Qt.AlignmentFlag.AlignCenter)  # Align headers (optional)

        # Connect the sectionMoved signal to update column indices
        header.sectionMoved.connect(self.update_column_indices)

    def update_column_indices(self, logical_index, old_visual_index, new_visual_index):
        """Handle the column reordering event."""
        print(f"Column moved: Logical index {logical_index} from position {old_visual_index} to {new_visual_index}")
        # Access the horizontal header
        header = self.uic.tableWidget.horizontalHeader()
        # Update the column order based on the header's current logical indices
        self.column_order = [header.logicalIndex(i) for i in range(header.count())]

    def open_excel_file(self):
        # Open file dialog to select Excel file
        file_path, _ = QFileDialog.getOpenFileName(
            self,
            "Chọn file Excel",
            "",
            "Excel Files (*.xlsx *.xlsm);;All Files (*)"
        )
        if not file_path:
            QMessageBox.warning(self, "Thông báo", "Không có file nào được chọn.")
            return

        try:
            workbook = load_workbook(file_path, data_only=True)

            # Logic for loading data into tables based on the number of sheets
            for sheet_index, sheet_name in enumerate(workbook.sheetnames):
                if sheet_index == 0:
                    table_widget = self.uic.tableWidget
                elif sheet_index == 1:
                    table_widget = self.uic.tableWidget_2
                elif sheet_index == 2:
                    table_widget = self.uic.tableWidget_3
                elif sheet_index == 3:
                    table_widget = self.uic.tableWidget_4
                elif sheet_index == 4:
                    table_widget = self.uic.tableWidget_5
                elif sheet_index == 5:
                    table_widget = self.uic.tableWidget_6
                else:
                    QMessageBox.warning(self, "Thông báo", "Chỉ hỗ trợ tối đa 6 sheet.")
                    break

                self.populate_table_with_sheet_data(table_widget, workbook[sheet_name])

                # Update the tab name to the sheet name
                self.uic.tabWidget.setTabText(sheet_index, sheet_name)

            QMessageBox.information(self, "Thông báo", f"Đã tải thành công dữ liệu từ file: {file_path}")
        except Exception as e:
            QMessageBox.critical(self, "Lỗi", f"Lỗi khi đọc file Excel: {e}")

    def populate_table_with_sheet_data(self, table_widget, sheet):
        table_widget.clearContents()

        """Populate the specified table widget with data from the sheet."""
        # Determine the actual number of columns with content
        max_col = 0
        for col_idx in range(1, sheet.max_column + 1):  # Iterate over columns
            if any(sheet.cell(row=row_idx, column=col_idx).value is not None for row_idx in range(1, sheet.max_row + 1)):
                max_col += 1

        # Determine the actual number of rows with content
        max_row = 0
        for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row, max_col=max_col):
            if any(cell.value is not None for cell in row):  # Check if any cell in the row has content
                max_row += 1

        # Set row and column counts
        table_widget.setRowCount(max_row)
        table_widget.setColumnCount(max_col)

        # Populate data
        for row in range(1, max_row + 1):
            for col in range(1, max_col + 1):
                cell_value = sheet.cell(row=row, column=col).value
                cell_value = str(cell_value) if cell_value is not None else ""
                table_widget.setItem(row - 1, col - 1, QTableWidgetItem(cell_value))

        # Add combo boxes to the column headers
        self.add_combobox_to_headers(table_widget, max_col)

    def add_combobox_to_headers(self, table_widget, column_count):
        """Add a combo box to each column header."""
        # Items for the combo boxes
        items = ['Tên sản phẩm', 'Model', 'Nhãn hiệu', 'Xuất xứ', 'Nhà sản xuất', 'Số lượng', 'Đơn vị tính']

        for col in range(column_count):
            # Create a QWidget to hold the combo box
            header_widget = QWidget()
            layout = QHBoxLayout()
            layout.setContentsMargins(0, 0, 0, 0)

            # Create and populate the combo box
            combo_box = QComboBox()
            combo_box.addItems(items)

            # Add combo box to the layout
            layout.addWidget(combo_box)
            header_widget.setLayout(layout)

            # Set the custom widget in the header
            table_widget.setCellWidget(0, col, header_widget)

    def merge_columns_with_same_combobox_text(self):
        """Merge text in all columns with the same combo box selection and keep only one column per group."""
        # Get the current tab index
        current_tab_index = self.uic.tabWidget.currentIndex()
        # Get the current tab's base widget
        tab_base_widget = self.uic.tabWidget.widget(current_tab_index)
        # Check if the base widget contains a QTableWidget
        table_widget = tab_base_widget.findChild(QTableWidget)

        if not isinstance(table_widget, QTableWidget):
            QMessageBox.warning(self, "Thông báo", "Không tìm thấy bảng dữ liệu trong tab hiện tại.")
            return

        max_row = table_widget.rowCount()
        max_col = table_widget.columnCount()

        # Get the current column order from the header
        header = table_widget.horizontalHeader()
        column_order = [header.logicalIndex(i) for i in range(header.count())]
        print("Current Column Order:", column_order)

        # Get the combobox selections for each column in the current order
        combo_box_values = {}
        for visual_index, col in enumerate(column_order):
            header_widget = table_widget.cellWidget(0, col)
            if header_widget and header_widget.layout():
                combo_box = header_widget.layout().itemAt(0).widget()
                if isinstance(combo_box, QComboBox):
                    combo_box_values[col] = combo_box.currentText()

        print("Tên cột (updated):", combo_box_values)

        # Group columns by their combo box text
        grouped_columns = {}
        for col, combo_value in combo_box_values.items():
            if combo_value not in grouped_columns:
                grouped_columns[combo_value] = []
            grouped_columns[combo_value].append(col)

        print("Grouped Columns:", grouped_columns)

        # Merge text for each group of columns based on the updated order
        for group, columns in grouped_columns.items():
            for row in range(max_row):
                # Concatenate text from all columns in the group
                merged_text = " ".join(
                    table_widget.item(row, col).text() if table_widget.item(row, col) else ""
                    for col in columns
                )
                # Update the first column in the group with the merged text
                if columns:
                    table_widget.setItem(row, columns[0], QTableWidgetItem(merged_text))

                # Remove all columns in the group except the first one
                for col in sorted(columns[1:], reverse=True):  # Reverse to avoid shifting indices
                    table_widget.setItem(row, col, QTableWidgetItem(""))
                    # table_widget.removeColumn(col)

        # Update the maximum column count after removing columns
        max_col = table_widget.columnCount()

        # Add a new column to the right
        new_col_index = max_col
        table_widget.insertColumn(new_col_index)
        table_widget.setHorizontalHeaderItem(new_col_index, QTableWidgetItem("Action"))

        # Add 'Find' buttons to every row in the new column
        for row in range(max_row):
            find_button = QPushButton("Find")
            find_button.clicked.connect(lambda _, r=row: self.find_product(r))
            table_widget.setCellWidget(row, new_col_index, find_button)

        QMessageBox.information(self, "Thông báo", "Đã gộp dữ liệu các cột có cùng tên và giữ lại một cột!")

    def find_product(self, row):
        """Print the item text in the row and the column with combobox text 'Tên sản phẩm'."""
        # Get the current tab index
        current_tab_index = self.uic.tabWidget.currentIndex()

        # Get the current tab's base widget
        tab_base_widget = self.uic.tabWidget.widget(current_tab_index)

        # Check if the base widget contains a QTableWidget
        table_widget = tab_base_widget.findChild(QTableWidget)

        if not isinstance(table_widget, QTableWidget):
            QMessageBox.warning(self, "Thông báo", "Không tìm thấy bảng dữ liệu trong tab hiện tại.")
            return

        # Find the column with combobox text 'Tên sản phẩm'
        target_column = None
        max_col = table_widget.columnCount()

        for col in range(max_col):
            header_widget = table_widget.cellWidget(0, col)
            if header_widget and header_widget.layout():
                combo_box = header_widget.layout().itemAt(0).widget()
                if isinstance(combo_box, QComboBox) and combo_box.currentText() == "Tên sản phẩm":
                    target_column = col
                    break

        if target_column is None:
            QMessageBox.warning(self, "Thông báo", "Không tìm thấy cột 'Tên sản phẩm'.")
            return

        # Get the item text in the specified row and target column
        item = table_widget.item(row, target_column)
        item_text = item.text() if item else "Không có dữ liệu"

        # Print the row and item text
        print(f"Row {row + 1}, Column {target_column + 1}: {item_text}")
        self.so_sanh(item_text)

    def so_sanh(self, text):
        ds = misc.sql_all("SELECT ten_san_pham FROM gia_tong_hop")
        print(ds)
        kq = self.find_most_similar(text, ds)
        print(f'Kết quả gần đúng nhất là: {kq[0]} - có điểm số: {kq[1]}.')

    def find_most_similar(self, target, string_list):
        """Find the most similar string to the target from a list."""
        most_similar = None
        highest_similarity = 0.0

        target = str(target).strip()  # Ensure target is a clean string
        for candidate in string_list:
            candidate = str(candidate).strip()  # Ensure candidate is a clean string
            # print(f"Comparing '{target}' with '{candidate}'")  # Debugging output
            similarity = SequenceMatcher(None, target, candidate).ratio()
            # print(f"Similarity: {similarity}")
            if similarity > highest_similarity:
                highest_similarity = similarity
                most_similar = candidate

        return most_similar, highest_similarity

    def delete_current_column(self):
        current_col = self.uic.tableWidget.currentColumn()
        if current_col != -1:
            self.uic.tableWidget.removeColumn(current_col)
        else:
            QMessageBox.warning(self, "Thông báo", "Hãy chọn một cột để xóa.")

    def delete_current_row(self):
        current_row = self.uic.tableWidget.currentRow()
        if current_row != -1:
            self.uic.tableWidget.removeRow(current_row)
        else:
            QMessageBox.warning(self, "Thông báo", "Hãy chọn một hàng để xóa.")

    def add_new_row(self):
        current_row = self.uic.tableWidget.currentRow()
        if current_row != -1:
            self.uic.tableWidget.insertRow(current_row + 1)
        else:
            QMessageBox.warning(self, "Thông báo", "Hãy chọn một hàng để thêm hàng mới.")

    def add_new_column(self):
        current_col = self.uic.tableWidget.currentColumn()
        if current_col != -1:
            self.uic.tableWidget.insertColumn(current_col)
            # Add a combo box to the new column header
            self.add_combobox_to_headers(self.uic.tableWidget.columnCount())
        else:
            QMessageBox.warning(self, "Thông báo", "Hãy chọn một cột để thêm cột mới.")


if __name__ == "__main__":
    app = QApplication(sys.argv)
    main_win = MainWindow()
    main_win.show()
    sys.exit(app.exec())
