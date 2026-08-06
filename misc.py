from openpyxl.drawing.image import Image
import os
import sys
import re
import unicodedata
import time
import threading
import mysql.connector
import mysql.connector.pooling
import requests
import datetime
import openpyxl
from PyQt6.QtWidgets import QFileDialog
import mysql.connector

db_config = {
    'host': 'db.fs.rsa.vn',
    'port': 3118,
    'user': 'firesmart',
    'password': '123@123a',
    'database': 'fs_mrb',
    'connection_timeout': 5,
    # TODO(S3): chuyển credential ra config ngoài repo — xem VIEC-CAN-LAM.md
}

# =====================================================================
#  CONNECTION POOL  (thêm 6/8/2026)
# ---------------------------------------------------------------------
#  TRƯỚC ĐÂY: mỗi câu lệnh SQL mở một kết nối MySQL mới rồi đóng ngay.
#  Toàn app có ~385 lượt gọi SQL ⇒ 385 lần bắt tay TCP + xác thực qua
#  Internet tới db.fs.rsa.vn. Một màn hình 20 query = 20 lần bắt tay.
#  Đây là nguyên nhân chính khiến app chậm.
#
#  BÂY GIỜ: giữ sẵn một pool kết nối và dùng lại. Lần gọi đầu tốn tiền
#  bắt tay, các lần sau gần như tức thì.
#
#  Chữ ký sql_all / sql_one / sql_commit GIỮ NGUYÊN — không file nào
#  phải sửa theo.
# =====================================================================

POOL_SIZE = 5          # đủ cho 1 app desktop; MySQL mặc định cho phép nhiều hơn
_pool = None
_pool_lock = threading.Lock()


def _build_pool():
    """Tạo pool. Lỗi thì trả None để rơi về kiểu kết nối trực tiếp."""
    try:
        return mysql.connector.pooling.MySQLConnectionPool(
            pool_name="fsales_pool",
            pool_size=POOL_SIZE,
            pool_reset_session=True,
            host=db_config['host'],
            port=db_config['port'],
            user=db_config['user'],
            password=db_config['password'],
            database=db_config['database'],
            connection_timeout=db_config.get("connection_timeout", 5),
            autocommit=False,
            use_pure=True,          # 🔥 BẮT BUỘC (giữ như cũ)
        )
    except Exception as e:
        print("Không tạo được connection pool, dùng kết nối trực tiếp:", e)
        return None


def _direct_connect():
    """Kết nối trực tiếp — đường lui khi pool hỏng hoặc cạn."""
    return mysql.connector.connect(
        host=db_config['host'],
        port=db_config['port'],
        user=db_config['user'],
        password=db_config['password'],
        database=db_config['database'],
        connection_timeout=db_config.get("connection_timeout", 5),
        use_pure=True,
    )


def _connect():
    """
    Lấy một kết nối. Ưu tiên pool; pool hỏng/cạn thì kết nối trực tiếp.

    Gọi .close() trên kết nối lấy từ pool KHÔNG đóng socket — nó trả
    kết nối về pool. Nên toàn bộ code cũ dùng `db.close()` vẫn đúng.
    """
    global _pool
    if _pool is None:
        with _pool_lock:
            if _pool is None:
                _pool = _build_pool()

    if _pool is not None:
        try:
            conn = _pool.get_connection()
            # Kết nối nằm không lâu có thể đã bị server ngắt (wait_timeout).
            # ping(reconnect=True) dựng lại mà không ném lỗi lên tầng trên.
            try:
                conn.ping(reconnect=True, attempts=2, delay=0)
            except Exception:
                pass
            return conn
        except Exception as e:
            print("Pool cạn hoặc lỗi, dùng kết nối trực tiếp:", e)

    return _direct_connect()


# ---------------------------------------------------------------------
#  PHÂN LOẠI LỖI
#  Trước đây MỌI lỗi đều bị thử lại 3 lần. Một câu SQL sai cú pháp cũng
#  bị chạy lại 3 lần ⇒ đơ giao diện ~10 giây rồi mới báo lỗi.
#  Giờ chỉ thử lại khi lỗi thật sự do KẾT NỐI.
# ---------------------------------------------------------------------

# Mã lỗi MySQL đáng thử lại: mất kết nối / server biến mất / hết chỗ
_RETRY_ERRNOS = {2002, 2003, 2006, 2013, 2055, 1040, 1053, 1205, 1213}


def _nen_thu_lai(e) -> bool:
    errno = getattr(e, "errno", None)
    if errno in _RETRY_ERRNOS:
        return True
    if isinstance(e, (mysql.connector.errors.InterfaceError,
                      mysql.connector.errors.OperationalError)):
        return True
    if isinstance(e, (mysql.connector.errors.ProgrammingError,
                      mysql.connector.errors.IntegrityError,
                      mysql.connector.errors.DataError)):
        return False          # lỗi do câu lệnh — thử lại cũng vô ích
    msg = str(e).lower()
    return any(k in msg for k in
               ("can't connect", "lost connection", "gone away",
                "timed out", "broken pipe", "not connected"))


def _dong(cur, db):
    """Đóng cursor + trả kết nối về pool. Không bao giờ ném lỗi."""
    try:
        if cur is not None:
            cur.close()
    except Exception:
        pass
    try:
        if db is not None:
            db.close()
    except Exception:
        pass


def sql_all(query, params=None, max_retry=3, default=None):
    retry = 0
    while True:
        db = cur = None
        try:
            db = _connect()
            cur = db.cursor()
            cur.execute(query, params)
            rows = cur.fetchall()
            return rows
        except Exception as e:
            if not _nen_thu_lai(e):
                print("sql_all error (không thử lại):", e)
                if default is not None:
                    return default
                raise
            retry += 1
            print(f"sql_all error (retry {retry}/{max_retry}):", e)
            if retry >= max_retry:
                if default is not None:
                    return default
                raise
            time.sleep(0.5)
        finally:
            _dong(cur, db)


def sql_one(query, params=None, max_retry=3, default=None):
    retry = 0
    while True:
        db = cur = None
        try:
            db = _connect()
            cur = db.cursor()
            cur.execute(query, params)
            row = cur.fetchone()
            # Phải đọc hết kết quả, nếu không kết nối trả về pool sẽ bẩn
            # và câu lệnh sau bị "Unread result found".
            try:
                cur.fetchall()
            except Exception:
                pass
            return row
        except Exception as e:
            if not _nen_thu_lai(e):
                print("sql_one error (không thử lại):", e)
                if default is not None:
                    return default
                raise
            retry += 1
            print(f"sql_one error (retry {retry}/{max_retry}):", e)
            if retry >= max_retry:
                if default is not None:
                    return default
                raise
            time.sleep(0.5)
        finally:
            _dong(cur, db)


def sql_commit(query, params=None, max_retry=3):
    retry = 0
    while True:
        db = cur = None
        try:
            db = _connect()
            cur = db.cursor()
            cur.execute(query, params)
            db.commit()
            return
        except Exception as e:
            msg = str(e)
            if "1062" in msg or "Duplicate entry" in msg:
                print("sql_commit integrity error:", e)
                raise

            try:
                if db is not None:
                    db.rollback()
            except Exception:
                pass

            if not _nen_thu_lai(e):
                print("sql_commit error (không thử lại):", e)
                raise

            retry += 1
            print(f"sql_commit error (retry {retry}/{max_retry}):", e)
            if retry >= max_retry:
                raise
            time.sleep(0.5)
        finally:
            _dong(cur, db)


def ensure_audit_schema():
    sql_commit(
        """
        CREATE TABLE IF NOT EXISTS sale_lead_audit_log (
            id BIGINT NOT NULL AUTO_INCREMENT,
            lead_id BIGINT NULL,
            log_line TEXT NOT NULL,
            created_at TIMESTAMP NOT NULL DEFAULT CURRENT_TIMESTAMP,
            PRIMARY KEY (id),
            KEY idx_lead_id (lead_id),
            KEY idx_created_at (created_at)
        ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4
        """,
        None,
    )


def audit_log(actor, action, field, old_value, new_value, lead_id=None):
    try:
        ensure_audit_schema()
        ts = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        old_txt = "(trống)" if old_value is None or str(old_value).strip() == "" else str(old_value)
        new_txt = "(trống)" if new_value is None or str(new_value).strip() == "" else str(new_value)
        line = f"[{ts}] {actor} | {action} | {field}: {old_txt} -> {new_txt}"
        sql_commit("INSERT INTO sale_lead_audit_log (lead_id, log_line) VALUES (%s, %s)", (lead_id, line))
    except Exception as e:
        print(f"audit_log error: {e}")


LEAD_TITLE_PLACEHOLDER_ANNA = 'Anna chờ sales đặt tên lead này'
LEAD_STATUS_ANNA_ASSIGNED = 'Anna đã giao việc'
LEAD_STATUS_ANNA_ASSIGNED_LEGACY = 'Đã giao việc từ Anna'
LEAD_READY_STATUS = {
    'Đã nhận việc',
    'Đã quá hạn báo giá',
    'Cần cập nhật lại',
    'Đã quá 10 ngày',
    'Đã báo giá',
    'Đã đặt hàng',
    'Đã thanh toán',
    'Đã giao hàng',
    'Đã trả lại toàn bộ',
    'Done - Thất bại',
}


def build_workflow_lead_title(name='', need=''):
    """Tạo tên cơ hội an toàn khi lead Anna vẫn đang dùng title placeholder.

    Lead do Anna tạo bắt đầu bằng placeholder để sales biết cần đặt tên lại.
    Tuy nhiên màn hình cập nhật lead hiện không có ô sửa `ten_co_hoi`, nên nếu
    sales đã nhận việc và bấm tạo báo giá thì cần tự materialize một tên cơ hội
    đủ dùng thay vì chặn workflow.
    """
    name = re.sub(r'\s+', ' ', str(name or '').strip())
    need = re.sub(r'\s+', ' ', str(need or '').strip())

    if name and need:
        title = f"{name} - {need}"
    elif need:
        title = need
    elif name:
        title = f"Nhu cầu báo giá - {name}"
    else:
        title = "Nhu cầu báo giá khách hàng"

    return title[:120].rstrip(' -')


def ensure_lead_title_before_quote(lead_id, actor='system'):
    row = sql_one("SELECT name, yc, ten_co_hoi FROM sale_lead WHERE lead_id = %s", (lead_id,))
    if not row:
        return

    old_title = str(row[2] or '').strip()
    if old_title and old_title != LEAD_TITLE_PLACEHOLDER_ANNA:
        return

    new_title = build_workflow_lead_title(row[0], row[1])
    sql_commit("UPDATE sale_lead SET ten_co_hoi = %s WHERE lead_id = %s", (new_title, lead_id))
    audit_log(actor or 'system', 'UPDATE_TITLE', 'ten_co_hoi', old_title or '(trống)', new_title, lead_id)


USER_BUSY_BLOCKING_STATUSES = {
    'Đã nhận việc',
    'Đã quá hạn báo giá',
    'Cần cập nhật lại',
    'Đã quá 10 ngày',
    LEAD_STATUS_ANNA_ASSIGNED,
    LEAD_STATUS_ANNA_ASSIGNED_LEGACY,
}

AUTO_ASSIGN_PRIMARY_USERS = ['Nguyễn Hải Hà', 'Lê Văn Việt']
AUTO_ASSIGN_FALLBACK_USERS = ['Hoàng Thị Thanh Nga', 'Nguyễn Ngọc Linh', 'Phí Ngọc Tùng']


def is_user_busy(user):
    row = sql_one("SELECT check_busy FROM user WHERE full_name = %s", (user,), default=(0,))
    try:
        return int((row or [0])[0] or 0) == 1
    except Exception:
        return False


def refresh_user_busy(user):
    user = str(user or '').strip()
    if not user or user == 'waiting':
        return 0

    row = sql_one(
        "SELECT COUNT(*) FROM sale_lead WHERE phu_trach = %s AND check_delete != '1' AND TRIM(status) IN (%s, %s, %s, %s, %s, %s)",
        (
            user,
            'Đã nhận việc',
            'Đã quá hạn báo giá',
            'Cần cập nhật lại',
            'Đã quá 10 ngày',
            LEAD_STATUS_ANNA_ASSIGNED,
            LEAD_STATUS_ANNA_ASSIGNED_LEGACY,
        ),
        default=(0,),
    )
    busy = 1 if int((row or [0])[0] or 0) > 0 else 0
    sql_commit("UPDATE user SET check_busy = %s WHERE full_name = %s", (busy, user))
    return busy


def _next_rotation_user(candidates):
    candidates = [str(x or '').strip() for x in (candidates or []) if str(x or '').strip()]
    if not candidates:
        return None

    rows = sql_all(
        "SELECT phu_trach, MAX(time_create) FROM sale_lead WHERE phu_trach IN ({}) GROUP BY phu_trach".format(','.join(['%s'] * len(candidates))),
        tuple(candidates),
        default=[],
    ) or []
    last_map = {str(r[0] or '').strip(): r[1] for r in rows}
    missing = [u for u in candidates if u not in last_map]
    if missing:
        return missing[0]
    return min(candidates, key=lambda u: last_map.get(u))


def pick_auto_assign_user(preferred_user=None):
    preferred_user = str(preferred_user or '').strip()
    if preferred_user and preferred_user != 'waiting' and not is_user_busy(preferred_user):
        return preferred_user

    primary_free = [u for u in AUTO_ASSIGN_PRIMARY_USERS if not is_user_busy(u)]
    if primary_free:
        return _next_rotation_user(primary_free) or primary_free[0]

    for user in AUTO_ASSIGN_FALLBACK_USERS:
        if not is_user_busy(user):
            return user

    return AUTO_ASSIGN_FALLBACK_USERS[-1]


def refresh_busy_for_lead(lead_id):
    row = sql_one("SELECT phu_trach FROM sale_lead WHERE lead_id = %s", (lead_id,), default=None)
    if row and row[0]:
        refresh_user_busy(row[0])


def refresh_busy_for_all_users_with_open_leads():
    rows = sql_all(
        "SELECT DISTINCT phu_trach FROM sale_lead WHERE phu_trach IS NOT NULL AND TRIM(phu_trach) != '' AND phu_trach != 'waiting' AND check_delete != '1'",
        None,
        default=[],
    ) or []
    for row in rows:
        refresh_user_busy(row[0])


def check_lead_ready_for_workflow(lead_id, *, allow_file_upload=False):
    row = sql_one("SELECT ten_co_hoi, status FROM sale_lead WHERE lead_id = %s", (lead_id,))
    if not row:
        return False, f"Không tìm thấy lead #{lead_id}"

    ten_co_hoi = (row[0] or '').strip()
    status = (row[1] or '').strip()

    if ten_co_hoi == LEAD_TITLE_PLACEHOLDER_ANNA:
        return False, "Lead chưa được đặt tên cơ hội. Vui lòng đổi tên trước khi thao tác tiếp."

    if allow_file_upload:
        return True, "OK"

    if status == LEAD_STATUS_ANNA_ASSIGNED:
        return False, "Lead đang ở trạng thái 'Anna đã giao việc'. Vui lòng chuyển sang 'Đã nhận việc' (hoặc trạng thái sau đó) để tiếp tục."

    if status not in LEAD_READY_STATUS:
        return False, "Lead chưa sẵn sàng xử lý. Vui lòng cập nhật trạng thái sang 'Đã nhận việc' (hoặc trạng thái sau đó)."

    return True, "OK"


def lookup_customer_profile(phone='', mst=''):
    phone = re.sub(r'\D', '', str(phone or ''))
    mst = re.sub(r'\D', '', str(mst or ''))

    profile = {
        'person_name': '',
        'person_phone': '',
        'company_name': '',
        'tax_code': '',
        'address': '',
        'email': '',
        'contact_name': '',
        'contact_phone': '',
        'contact_title': '',
        'contact_email': '',
        'latest_lead_title': '',
        'latest_owner': '',
        'lead_count': 0,
        'won_count': 0,
        'history_note': '',
        'company_found': False,
        'person_found': False,
    }

    row = None
    leads = []
    won = []
    company = None

    try:
        if phone and len(phone) == 10:
            row = sql_one(
                "SELECT ten, dien_thoai, leads, ten_cong_ty, mst_cong_ty, address FROM ds_ca_nhan WHERE dien_thoai = %s",
                (phone,),
                default=None
            )
            if row:
                profile['person_found'] = True
                profile['person_name'] = str(row[0] or '')
                profile['person_phone'] = str(row[1] or '')
                profile['company_name'] = str(row[3] or '')
                profile['tax_code'] = re.sub(r'\D', '', str(row[4] or ''))
                profile['address'] = str(row[5] or '')
                if not mst and profile['tax_code']:
                    mst = profile['tax_code']

            leads = sql_all(
                "SELECT lead_id, name, sdt, company, mst, yc, address, phu_trach, ten_co_hoi "
                "FROM sale_lead WHERE sdt = %s ORDER BY lead_id DESC",
                (phone,),
                default=[]
            ) or []
            if leads:
                profile['lead_count'] = len(leads)
                latest = leads[0]
                profile['latest_owner'] = str(latest[7] or '')
                profile['latest_lead_title'] = str(latest[8] or '')
                if not profile['person_name']:
                    profile['person_name'] = str(latest[1] or '')
                if not profile['company_name']:
                    profile['company_name'] = str(latest[3] or '')
                if not profile['tax_code']:
                    profile['tax_code'] = re.sub(r'\D', '', str(latest[4] or ''))
                if not profile['address']:
                    profile['address'] = str(latest[6] or '')
                if not mst and profile['tax_code']:
                    mst = profile['tax_code']

                won = sql_all("SELECT lead_id FROM sale_lead WHERE sdt = %s AND dat_hang = 'T'", (phone,), default=[]) or []
                profile['won_count'] = len(won)
                history_note = f"Đã từng có {len(leads)} cơ hội bán hàng"
                if won:
                    history_note += f" và có {len(won)} đơn hàng thành công."
                owners = sorted(set([str(ele[7]) for ele in leads if len(ele) > 7 and ele[7]]))
                if owners:
                    history_note += " Người đã từng liên hệ: " + ", ".join(owners) + "."
                profile['history_note'] = history_note

        if mst:
            company = sql_one(
                "SELECT ten_cong_ty, dia_chi, mst, nguoi_lien_he, sdt_nguoi_lh, chuc_vu_nlh, email_nlh, email_cong_ty, dien_thoai_cong_ty "
                "FROM ds_cong_ty WHERE mst = %s",
                (mst,),
                default=None
            )
            if company:
                profile['company_found'] = True
                profile['company_name'] = str(company[0] or profile['company_name'])
                profile['address'] = str(company[1] or profile['address'])
                profile['tax_code'] = re.sub(r'\D', '', str(company[2] or profile['tax_code']))
                profile['contact_name'] = str(company[3] or '')
                profile['contact_phone'] = str(company[4] or '')
                profile['contact_title'] = str(company[5] or '')
                profile['contact_email'] = str(company[6] or '')
                profile['email'] = str(company[6] or company[7] or profile['email'])
                if not profile['person_name']:
                    profile['person_name'] = str(company[3] or '')
                if not profile['person_phone']:
                    profile['person_phone'] = str(company[4] or company[8] or '')

                if profile['history_note']:
                    profile['history_note'] += f"\nCông ty {profile['company_name']} đã từng có lịch sử giao dịch."
                else:
                    profile['history_note'] = f"Công ty {profile['company_name']} đã từng có lịch sử giao dịch."
    except Exception as e:
        print(f'lookup_customer_profile error: {e}')
        profile['history_note'] = (profile.get('history_note') + '\n' if profile.get('history_note') else '') + 'Không thể tra cứu thêm từ DB lúc này.'

    return profile


def tao_bao_gia(lead_id, user):

    now = datetime.datetime.now()

    # Xóa báo giá cũ không có nội dung
    kq = sql_all("SELECT * FROM ds_bao_gia WHERE noi_dung IS NULL", None)
    for item in kq:
        item_date = datetime.datetime.strptime(item[2], '%d/%m/%y').date()
        if item_date < now.date():
            sql_commit("DELETE FROM ds_bao_gia WHERE so_bg = %s", (item[0],))

    # Tạo số báo giá mới
    kq = sql_one("SELECT MAX(so_bg) FROM ds_bao_gia")
    sobaogia = 1 if kq[0] is None else int(kq[0]) + 1

    lead_id = int(lead_id)

    ensure_lead_title_before_quote(lead_id, user)

    ok, msg = check_lead_ready_for_workflow(lead_id)
    if not ok:
        raise ValueError(msg)

    d = now.strftime("%d/%m/%y")

    sql_commit(
        "INSERT INTO ds_bao_gia (so_bg, lead_id, ngaythang, sotien, user, tieu_de) VALUES (%s, %s, %s, %s, %s, %s)",
        (sobaogia, lead_id, d, 0, user, 'Cần đổi tên của báo giá này!')
    )
    audit_log(user, 'CREATE_QUOTE', 'so_bg', '-', sobaogia, lead_id)

    return 'Thông tin hợp lệ – báo giá đã được tạo', sobaogia


def send_to_telegram(message):

    apiToken = '6469906602:AAEuTZ3y0tOMug8qX8DKoRdOLZN1bycb-WA'
    chatID = '-928385747'
    apiURL = f'https://api.telegram.org/bot{apiToken}/sendMessage'

    try:
        response = requests.post(apiURL, json={'chat_id': chatID, 'text': message})
    except Exception as e:
        print(e)


def get_resource_path(relative_path):
    """
    Dùng được cả khi chạy Python thường và khi build PyInstaller (.exe)
    """
    if hasattr(sys, "_MEIPASS"):
        return os.path.join(sys._MEIPASS, relative_path)
    return os.path.join(os.path.abspath("."), relative_path)


def insert_logo_by_company(sheet, cty):
    """
    Chèn logo vào sheet theo combo công ty
    """
    LOGO_MAP = {
        "BG PCCC.vn": "assets/logo_pccc.png",
        "BG Infutech": "assets/logo_infutech.png",
        "BG Bach Khoa": "assets/logo_bach_khoa.png",
    }

    logo_rel_path = LOGO_MAP.get(cty)

    if not logo_rel_path:
        print(f"⚠️ Không tìm thấy logo cho công ty: {cty}")
        return

    logo_path = get_resource_path(logo_rel_path)

    if not os.path.exists(logo_path):
        print(f"❌ File logo không tồn tại: {logo_path}")
        return

    img = Image(logo_path)

    # 🎯 Chỉnh kích thước chuẩn cho báo giá
    img.width = 180
    img.height = 70

    # 🎯 Vị trí đặt logo (bạn có thể đổi)
    sheet.add_image(img, "A1")


def save_excel(so_bg, user, phone, cty, thue=None):

    # Get lead_id
    kq = sql_one("SELECT lead_id, tieu_de FROM ds_bao_gia WHERE so_bg = %s", (so_bg,))
    lead_id = str(kq[0])
    tieu_de = kq[1]
    # Step 1: Get address
    ttkh = sql_one("SELECT company, name, address, sdt, mst FROM sale_lead WHERE lead_id = %s", (lead_id,)) or ''

    # Step 2: Get quote data
    kq = sql_one("SELECT so_bg, ngaythang, sotien, noi_dung, ghi_chu, user, dien_thoai, sum8, sum10, sum0 FROM ds_bao_gia WHERE so_bg = %s", (so_bg,))
    goods = [item.split('|') for item in kq[3].split('@')]

    # Step 3: Select template
    if thue == 'Thue':
        template_path = get_resource_path('bao_gia_thue.xlsx')
    else:
        if cty == 'BG PCCC.vn':
            template_path = get_resource_path('bao_gia_mau.xlsx')
        elif cty == 'BG Infutech':
            template_path = get_resource_path('bao_gia_mau_infutech.xlsx')
        else:
            template_path = get_resource_path('bao_gia_mau_bach_khoa.xlsx')

    if not os.path.exists(template_path):
        return f"Không tìm thấy file mẫu báo giá: {template_path}"

    workbook = openpyxl.load_workbook(template_path)

    sheet = workbook['Quotation']
    #insert_logo_by_company(sheet, cty)

    now_str = datetime.datetime.now().strftime('%d-%m-%Y')

    # Step 4: Fill general info
    sheet.cell(row=5, column=2).value = f'Kính gửi: {ttkh[0]}'
    sheet.cell(row=6, column=2).value = f'Anh/Chị: {ttkh[1]}'
    sheet.cell(row=7, column=2).value = f'Địa chỉ: {ttkh[2]}'
    sheet.cell(row=8, column=2).value = f'Điện thoại: {ttkh[3]}'
    sheet.cell(row=5, column=9).value = now_str
    sheet.cell(row=6, column=9).value = f'BG-{so_bg}-{lead_id}'
    sheet.cell(row=9, column=2).value = f"V/v: {tieu_de}"

    # Step 5: Fill goods list
    for i, item in enumerate(goods):
        try:
            stt = str(i + 1)
            tax_marker = item[6] if len(item) > 6 else '0'
            prefix = {'0': '', '8': '.', '10': '..'}.get(tax_marker, '...')
            item.insert(0, f'{stt}{prefix}')
            total = int(item[5]) * int(item[6].replace(',', ''))
            item.append(str(total))

            for col in range(2, 9):
                cell = sheet.cell(row=12 + i, column=col)
                cell.value = int(item[col - 2]) if col == 8 else item[col - 2]
                if col == 8:
                    cell.number_format = "#,##0"
        except Exception as e:
            print("Lỗi khi điền dữ liệu hàng hoá:", e)

    # Step 6: Delete extra rows
    sheet.delete_rows(13 + len(goods), 80 - len(goods))

    # Step 7: Tính thuế, tổng tiền
    sheet.cell(row=13 + len(goods), column=9).value = f'=SUM(I12:I{11 + len(goods)})'

    # LUÔN TÍNH VAT – an toàn kiểu dữ liệu
    sum8 = float(kq[7] or 0)
    sum10 = float(kq[8] or 0)

    sheet.cell(row=14 + len(goods), column=9).value = round(sum8 * 0.08)
    sheet.cell(row=15 + len(goods), column=9).value = round(sum10 * 0.1)

    if thue == 'Thue':
        sheet.cell(row=15 + len(goods), column=9).value = sum(int(item[5]) for item in goods) * 100000
    sheet.cell(row=16 + len(goods), column=9).value = f'=SUM(I{13 + len(goods)}:I{15 + len(goods)})'

    # Step 8: Tên người làm, sĐT
    sheet.cell(row=22 + len(goods), column=8).value = user
    sheet.cell(row=23 + len(goods), column=8).value = phone

    # Step 9: Save file
    # Step 9: Save file
    #root = Tk()
    #root.withdraw()


    def clean_filename(name):

        # Chuẩn hóa unicode
        name = unicodedata.normalize('NFKD', name)

        # Loại bỏ ký tự không phải ASCII in được
        name = ''.join(c for c in name if c.isprintable())

        # Thay mọi khoảng trắng đặc biệt bằng space thường
        name = re.sub(r'\s+', ' ', name)

        # Loại bỏ ký tự Windows cấm
        name = re.sub(r'[\\/*?:"<>|]', "", name)

        # Cắt khoảng trắng đầu/cuối
        name = name.strip()

        # Giới hạn độ dài
        return name[:120]

    raw_name = f"BG {so_bg} - {ttkh[1]} - {tieu_de}"
    safe_name = clean_filename(raw_name)

    default_file_name = f"{safe_name}.xlsx"

    # Ask the user to select a location to save the modified file
    #file_save_path = filedialog.asksaveasfilename(
     #   initialfile=default_file_name,
      #  defaultextension=".xlsx",
       # filetypes=[("Excel Files", "*.xlsx")]
    #)

    file_save_path, _ = QFileDialog.getSaveFileName(
        None,
        "Save Excel File",
        default_file_name,
        "Excel Files (*.xlsx)"
    )

    if not file_save_path:
        return 'Chưa SAVE file excel.'

    try:
        workbook.save(file_save_path)
        tenfile = file_save_path.split('/')[-1]
        return "Đã SAVE file:   " + tenfile
    except Exception as e:
        print(e)
        return 'Chưa save file, kiểm tra xem có file nào trùng tên đang mở không!'

    # Check if the user canceled the save dialog
    if not file_save_path:
        tex = 'Chưa SAVE file excel.'
    else:
        try:
            # Save the modified workbook to the selected location
            workbook.save(file_save_path)

            # Find the last occurrence of '/'
            last_slash = file_save_path.rfind('/')

            # Extract substring from last slash to the end
            if last_slash != -1:  # Check if '/' is found
                tenfile = file_save_path[last_slash + 1:]  # +1 to exclude the slash itself
            else:
                tenfile = file_save_path  # If '/' is not found, return the original string

            tex = "Đã SAVE file:   " + tenfile

            check = 1
        except Exception as e:
            print(e)
            check = 0
            tex = 'Chưa save file, kiểm tra xem có file nào trùng tên đang mở không!'
    # Close the tkinter root window
    root.destroy()
    return tex


def header_label(user):
    # Dùng sql_one để đi qua pool + cơ chế retry chung (sửa 6/8/2026).
    row = sql_one("""
        SELECT COUNT(*) FROM sale_lead
        WHERE YEAR(time_create) = YEAR(CURDATE())
        AND MONTH(time_create) = MONTH(CURDATE())
        AND nguoi_tao_lead = %s
    """, (user,), default=(0,))
    result = row[0] if row else 0
    return f"Tháng này bạn đã tạo ra {result} cơ hội."


def header_label_doanh_so(user):
    # Ghi lịch sử tạo lead của user
    result = sql_all("SELECT * FROM ds_don_hang WHERE nguoi_tao = %s", (user,))
    txt = 'Đã chốt ' + str(len(result)) + ' đơn hàng, tổng giá trị ' + "{:,}".format(
        sum(item[2] for item in result)) + ' VNĐ.'

    return txt


def save_price_list(price_list):
    saved = False
    while not saved:
        try:
            for e in price_list:
                code = ("UPDATE gia_tong_hop SET ten_san_pham = %s, nhan_hieu = %s, xuat_xu = %s, don_vi = %s, "
                        "gia_cap_1 = %s, gia_cap_2 = %s, gia_ban_le = %s, vat = %s, gia_dau_vao = %s WHERE model = %s")
                val = (e[0], e[2], e[3], e[4], e[5], e[6], e[7], e[8], e[9], e[1])
                sql_commit(code, val)
                time.sleep(0.05)
                print('Đã lưu đến sản phẩm: ', e[0])
            saved = True
        except Exception as e1:
            saved = False
            time.sleep(0.5)
            print(e1)


# =====================================================================
# KHẢO SÁT HIỆN TRƯỜNG — sale_lead_khao_sat
# =====================================================================
KHAO_SAT_FIELDS = [
    # ===== A. Thông tin chung công trình + cơ sở (v5 mở rộng PC01) =====
    "kh_cty", "kh_mst", "ten_cong_trinh", "dia_chi", "cong_nang_k",
    "nam_hoat_dong", "nganh_nghe",
    "hinh_thuc_so_huu", "trang_thai", "thanh_phan_kt",
    "co_quan_cap_tren", "nguoi_quan_ly", "quan_ly_sdt",
    "dai_dien_pl_ten", "dai_dien_pl_sdt",
    "ct_doc_lap", "thuoc_dm_nguyhiem", "thuoc_dm_thamduyet",
    # ===== B. Quy mô + Kỹ thuật + Giao thông CC =====
    "dt_san_tong", "dt_xay_dung", "cao_pccc",
    "so_tang_noi", "so_tang_ham", "so_phong",
    "dai_nha", "rong_nha", "kc_ct_ke_ben",
    "bac_chiu_lua", "cap_nhc", "hang_nguy_hiem",
    "so_nguoi_du_kien", "so_chau",
    "dgt_rong", "dgt_cao", "bai_do_xe_cc", "xe_cc_tiep_can",
    # ===== C. Hạ tầng + Hệ thống PCCC sẵn có =====
    "nguon_nuoc", "nguon_nuoc_json", "nguon_nuoc_chi_tiet",
    "so_be_cc", "khoi_tich_be", "vi_tri_be",
    "so_tru_cc", "vi_tri_tru",
    "nguon_dien", "co_may_phat",
    "mp_cong_suat_kva", "mp_thoi_gian_chay_h",
    "truyen_tin_da_lap", "co_xe_chua_chay", "phuong_tien_cc_text",
    "he_thong_sn_json",
    # ===== D. Pháp lý + Văn bản + Tài liệu =====
    "vb_thamduyet_so", "vb_thamduyet_ngay", "vb_thamduyet_cq",
    "vb_nghiemthu_so", "vb_nghiemthu_ngay", "vb_nghiemthu_cq",
    "bh_cong_ty", "bh_so_hd", "bh_ngay_het_han",
    "hd_bao_duong", "hd_bao_duong_ncc",
    "tai_lieu_json", "lich_su_pccc", "bkl_drive_info",
    # ===== E. Thương mại + Đánh giá sales =====
    "yc_kh", "ngan_sach", "deadline",
    "nguoi_quyet_dinh", "qd_ten", "qd_chuc_vu", "qd_sdt",
    "lien_he_ky_thuat", "lh_ten", "lh_chuc_vu", "lh_sdt",
    "doi_thu_da_bao_gia", "doi_thu_ten",
    "danh_gia_sales", "buoc_tiep_theo",
    # ===== F. Khảo sát + Đội PCCC cơ sở =====
    "ngay_khao_sat", "nguoi_khao_sat",
    "nguoi_tiep_don", "td_ten", "td_chuc_vu", "td_sdt",
    "doi_tong_doi_vien", "doi_truong_ten", "doi_truong_sdt",
    "doi_tong_doi_vien", "doi_truong_ten", "doi_truong_sdt",
    "so_nguoi_pccc",
    # ===== G. Khối nhà + Khu vực ngoài nhà (JSON snapshot) =====
    "khoi_nha_json", "khu_vuc_json",
]


def doc_khao_sat(lead_id):
    """Đọc khảo sát của 1 lead. Trả dict hoặc {} nếu chưa có."""
    cols = ", ".join(KHAO_SAT_FIELDS)
    row = sql_one(
        f"SELECT {cols} FROM sale_lead_khao_sat WHERE lead_id = %s",
        (lead_id,))
    if not row:
        return {}
    return dict(zip(KHAO_SAT_FIELDS, row))


def luu_khao_sat(lead_id, data: dict):
    """Upsert khảo sát (INSERT ON DUPLICATE KEY UPDATE)."""
    cols = ["lead_id"] + KHAO_SAT_FIELDS
    placeholders = ", ".join(["%s"] * len(cols))
    updates = ", ".join(f"{c}=VALUES({c})" for c in KHAO_SAT_FIELDS)
    values = [lead_id] + [data.get(f) for f in KHAO_SAT_FIELDS]
    sql_commit(
        f"INSERT INTO sale_lead_khao_sat ({', '.join(cols)}) "
        f"VALUES ({placeholders}) "
        f"ON DUPLICATE KEY UPDATE {updates}",
        values)
